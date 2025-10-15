import streamlit as st
import pandas as pd
import openpyxl
import tempfile

st.set_page_config(page_title="Lấy Kê Giờ GV", layout="wide")
st.title("Lấy dữ liệu kê giờ từ file GV sang file tổng hợp Khoa")


st.header("Bước 1: Upload file của Giáo viên")
uploaded_gv_file = st.file_uploader("Tải lên file Excel của Giáo viên (xls/xlsx)", type=["xls", "xlsx"], key="gv_file")

st.header("Bước 2: Upload file tổng hợp Khoa")
uploaded_khoa_file = st.file_uploader("Tải lên file Excel tổng hợp Khoa (xls/xlsx)", type=["xls", "xlsx"], key="khoa_file")

gv_path = None
khoa_path = None
if uploaded_gv_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_gv:
        tmp_gv.write(uploaded_gv_file.read())
        gv_path = tmp_gv.name
if uploaded_khoa_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_khoa:
        tmp_khoa.write(uploaded_khoa_file.read())
        khoa_path = tmp_khoa.name

if uploaded_gv_file:
    # Hiển thị danh sách sheet của file GV
    try:
        wb_gv_sheet = openpyxl.load_workbook(gv_path, data_only=True)
        sheet_list = wb_gv_sheet.sheetnames
        st.info(f"Các sheet trong file GV: {sheet_list}")
        if "Ke_gio_HK2_Cả_năm" not in sheet_list:
            st.warning("Không tìm thấy sheet 'Ke_gio_HK2_Cả_năm' trong file GV. Vui lòng kiểm tra lại tên sheet hoặc file.")
    except Exception as e:
        st.error(f"Lỗi khi đọc sheet của file GV: {e}")
    st.header("Bước 3: Trích xuất hoạt động quy đổi từ file tổng hợp Khoa (sheet CAC_HOAT_DONG)")
    hoatdongquydoi = []
    ten_hd_list = []
    so_col_list = []
    if khoa_path:
        # Khởi tạo ws_hd trước khi lấy danh sách giáo viên
        wb_khoa_hd = openpyxl.load_workbook(khoa_path, data_only=True)
        if "CAC_HOAT_DONG" in wb_khoa_hd.sheetnames:
            ws_hd = wb_khoa_hd["CAC_HOAT_DONG"]
            giaovien_list = []
            for row in range(8, ws_hd.max_row + 1):
                ten_gv = ws_hd.cell(row=row, column=2).value
                if ten_gv is not None and str(ten_gv).strip() != "":
                    giaovien_list.append(str(ten_gv).strip())
            st.subheader("Chọn giáo viên để cập nhật dữ liệu")
            selected_gv = st.selectbox("Giáo viên", giaovien_list, key="chon_gv")
        else:
            ws_hd = None
            st.warning("Không tìm thấy sheet 'CAC_HOAT_DONG' trong file tổng hợp Khoa.")
        try:
            wb_khoa_hd = openpyxl.load_workbook(khoa_path, data_only=True)
            if "CAC_HOAT_DONG" in wb_khoa_hd.sheetnames:
                ws_hd = wb_khoa_hd["CAC_HOAT_DONG"]
                # Lấy các giá trị từ cột E (5) đến cột AE (31) ở dòng 7
                for col in range(5, 32):
                    ten_hd = ws_hd.cell(row=7, column=col).value
                    if ten_hd is not None and str(ten_hd).strip() != "":
                        hoatdongquydoi.append({"ten_hd": str(ten_hd), "so_col": col})
                        ten_hd_list.append(str(ten_hd))
                        so_col_list.append(col)
            else:
                st.warning("Không tìm thấy sheet 'CAC_HOAT_DONG' trong file tổng hợp Khoa.")
        except Exception as e:
            st.error(f"Lỗi khi đọc sheet CAC_HOAT_DONG: {e}")
    sheet_name = "Ke_gio_HK2_Cả_năm"
    try:
        wb_gv = openpyxl.load_workbook(gv_path, data_only=True)
        if sheet_name in wb_gv.sheetnames:
            ws_gv = wb_gv[sheet_name]
            # Tìm vị trí dòng 'CÁC HOẠT ĐỘNG KHÁC QUY RA GIỜ CHUẨN' ở cột A
            start_row = None
            for row in range(1, ws_gv.max_row + 1):
                val = str(ws_gv.cell(row=row, column=1).value).strip() if ws_gv.cell(row=row, column=1).value is not None else ''
                if val.upper() == 'CÁC HOẠT ĐỘNG KHÁC QUY RA GIỜ CHUẨN':
                    start_row = row
                    break
            # Sau khi gặp dòng này, tìm dòng TT ở cột A phía dưới
            tt_row = None
            if start_row is not None:
                for row in range(start_row+1, ws_gv.max_row + 1):
                    val = str(ws_gv.cell(row=row, column=1).value).strip() if ws_gv.cell(row=row, column=1).value is not None else ''
                    if val == 'TT':
                        tt_row = row
                        break
            # Nếu tìm thấy dòng TT, lấy dữ liệu từ dòng đó trở xuống đến khi gặp dòng trống ở cột A
            data_rows = []
            if tt_row is not None:
                for row in range(tt_row, ws_gv.max_row + 1):
                    val_a = ws_gv.cell(row=row, column=1).value
                    if val_a is None or str(val_a).strip() == '':
                        break
                    # Lấy các cột: TT (A), Nội dung (B), Tên/Tiêu đề hoạt động (D), Quy ra giờ (L)
                    row_data = [ws_gv.cell(row=row, column=1).value,  # TT
                                ws_gv.cell(row=row, column=2).value,  # Nội dung
                                ws_gv.cell(row=row, column=4).value,  # Tên/Tiêu đề hoạt động
                                ws_gv.cell(row=row, column=12).value] # Quy ra giờ
                    # Nếu toàn bộ row_data đều trống thì bỏ qua
                    if all(x is None or str(x).strip() == '' for x in row_data):
                        continue
                    data_rows.append(row_data)
            # Tạo DataFrame và đặt tên cột
            df_result = None
            if data_rows:
                df_result = pd.DataFrame(data_rows, columns=["TT", "Nội dung", "Tên/Tiêu đề hoạt động (Số QĐ ban hành/...)", "Quy ra giờ"])
                # Xóa các dòng trống hoàn toàn
                df_result.dropna(how='all', inplace=True)
                # Xóa các dòng mà cả 3 cột nội dung đều trống
                df_result = df_result.dropna(subset=["Nội dung", "Tên/Tiêu đề hoạt động (Số QĐ ban hành/...)", "Quy ra giờ"], how='all')
                # Loại bỏ dòng tiêu đề (TT, NỘI DUNG, Tên..., Quy ra giờ)
                df_result = df_result[~((df_result["TT"] == "TT") & (df_result["Nội dung"] == "NỘI DUNG"))]
                df_result = df_result[~((df_result["TT"] == "TT") & (df_result["Tên/Tiêu đề hoạt động (Số QĐ ban hành/...)"] == "Tên/Tiêu đề hoạt động (Số QĐ ban hành/...)") & (df_result["Quy ra giờ"] == "Quy ra giờ"))]
                st.subheader("Dữ liệu hoạt động khác quy ra giờ chuẩn")
                st.dataframe(df_result)
                st.markdown("---")
                st.subheader("Chọn hoặc chỉnh sửa từng dòng hoạt động")
                selected_rows = []
                for idx, row in df_result.iterrows():
                    col1, col2, col3, col4 = st.columns([1,4,4,1])
                    with col1:
                        tt_val = st.text_input("TT", value=str(row["TT"]), key=f"tt_{idx}")
                    with col2:
                        nd_val = st.selectbox("Nội dung", ten_hd_list, index=ten_hd_list.index(str(row["Nội dung"])) if str(row["Nội dung"]) in ten_hd_list else 0, key=f"nd_{idx}")
                    with col3:
                        ten_val = st.text_input("Tiêu đề hoạt động", value=str(row["Tên/Tiêu đề hoạt động (Số QĐ ban hành/...)"]), key=f"ten_{idx}")
                    with col4:
                        qg_val = st.text_input("Quy ra giờ", value=str(row["Quy ra giờ"]), key=f"qg_{idx}")
                    selected_rows.append({
                        "TT": tt_val,
                        "Nội dung": nd_val,
                        "Tên/Tiêu đề hoạt động (Số QĐ ban hành/...)": ten_val,
                        "Quy ra giờ": qg_val
                    })
                st.markdown("---")
                if st.button("Cập nhật dữ liệu GV vào bảng tổng hợp Khoa", key="btn_update_gv"):
                    if khoa_path and len(selected_rows) > 0 and 'selected_gv' in locals():
                        wb = openpyxl.load_workbook(khoa_path)
                        ws_hd = wb["CAC_HOAT_DONG"] if "CAC_HOAT_DONG" in wb.sheetnames else wb.active
                        # Tìm hàng theo tên GV đã chọn trong cột B
                        row_gv = None
                        for row in range(8, ws_hd.max_row + 1):
                            cell_val = str(ws_hd.cell(row=row, column=2).value).strip() if ws_hd.cell(row=row, column=2).value else ""
                            if cell_val == selected_gv:
                                row_gv = row
                                break
                        if row_gv is not None:
                            # Ghi dữ liệu Quy ra giờ vào cột tương ứng với hoạt động quy đổi
                            for r in selected_rows:
                                ten_hd = r["Nội dung"]
                                qg_val = r["Quy ra giờ"]
                                # Tìm số cột tương ứng với ten_hd
                                so_col = None
                                for hd in hoatdongquydoi:
                                    if hd["ten_hd"] == ten_hd:
                                        so_col = hd["so_col"]
                                        break
                                if so_col is not None and qg_val not in [None, "", "nan"]:
                                    try:
                                        ws_hd.cell(row=row_gv, column=so_col).value = float(qg_val)
                                    except:
                                        ws_hd.cell(row=row_gv, column=so_col).value = qg_val
                            wb.save(khoa_path)
                            st.success(f"Đã cập nhật dữ liệu hoạt động quy đổi cho giáo viên '{selected_gv}' vào bảng tổng hợp Khoa. Bạn có thể tiếp tục chọn giáo viên khác để cập nhật tiếp.")
                            # Hiển thị lại dữ liệu sheet CAC_HOAT_DONG sau khi đã cập nhật
                            try:
                                df_khoa_updated = pd.read_excel(khoa_path, sheet_name="CAC_HOAT_DONG", header=6, engine='openpyxl')
                                # Đặt lại tên cột bằng dòng đầu tiên, loại bỏ dòng này khỏi dữ liệu
                                df_khoa_updated.columns = df_khoa_updated.iloc[0].fillna("")
                                df_khoa_updated = df_khoa_updated[1:].reset_index(drop=True)
                                df_khoa_updated = df_khoa_updated.where(pd.notnull(df_khoa_updated), "")
                                st.subheader(f"Xem trước dữ liệu sheet CAC_HOAT_DONG sau khi cập nhật cho giáo viên '{selected_gv}'")
                                st.dataframe(df_khoa_updated)
                            except Exception as e:
                                st.error(f"Lỗi khi đọc lại sheet CAC_HOAT_DONG sau cập nhật: {e}")
                        else:
                            st.error(f"Không tìm thấy tên giáo viên '{selected_gv}' trong cột B của sheet CAC_HOAT_DONG.")
            else:
                st.info("Không tìm thấy dữ liệu phù hợp trong sheet hoặc file GV.")
        else:
            st.warning(f"Không tìm thấy sheet '{sheet_name}' trong file GV.")
    except Exception as e:
        st.error(f"Lỗi khi trích xuất dữ liệu hoạt động khác: {e}")
    st.header("Bước 4: Lấy dữ liệu từ file GV và truyền vào file tổng hợp Khoa")
    # gv_path và khoa_path đã được tạo ở trên, không cần tạo lại hoặc ghi đè
    # Đọc dữ liệu từ file GV
    # Đã bỏ phần đọc và hiển thị dữ liệu file GV bằng pandas
    # Đọc dữ liệu từ file tổng hợp Khoa
    try:
        df_khoa = pd.read_excel(khoa_path, engine='openpyxl')
        st.subheader("Xem trước dữ liệu file tổng hợp Khoa")
        st.dataframe(df_khoa)
    except Exception as e:
        st.error(f"Lỗi đọc file tổng hợp Khoa: {e}")
        df_khoa = None
    # Thực hiện truyền dữ liệu (ví dụ: cập nhật theo tên GV)
    if khoa_path and len(selected_rows) > 0 and 'selected_gv' in locals():
        st.info("Thực hiện truyền dữ liệu từ file GV sang file tổng hợp Khoa theo giáo viên đã chọn và hoạt động quy đổi.")
        wb = openpyxl.load_workbook(khoa_path)
        ws_hd = wb["CAC_HOAT_DONG"] if "CAC_HOAT_DONG" in wb.sheetnames else wb.active
        # Tìm hàng theo tên GV đã chọn trong cột B
        row_gv = None
        for row in range(8, ws_hd.max_row + 1):
            cell_val = str(ws_hd.cell(row=row, column=2).value).strip() if ws_hd.cell(row=row, column=2).value else ""
            if cell_val == selected_gv:
                row_gv = row
                break
        if row_gv is not None:
            # Ghi dữ liệu Quy ra giờ vào cột tương ứng với hoạt động quy đổi
            for r in selected_rows:
                ten_hd = r["Nội dung"]
                qg_val = r["Quy ra giờ"]
                # Tìm số cột tương ứng với ten_hd
                so_col = None
                for hd in hoatdongquydoi:
                    if hd["ten_hd"] == ten_hd:
                        so_col = hd["so_col"]
                        break
                if so_col is not None and qg_val not in [None, "", "nan"]:
                    try:
                        ws_hd.cell(row=row_gv, column=so_col).value = float(qg_val)
                    except:
                        ws_hd.cell(row=row_gv, column=so_col).value = qg_val
            wb.save(khoa_path)
            with open(khoa_path, 'rb') as f:
                st.download_button(
                    label="Tải xuống file tổng hợp Khoa đã cập nhật",
                    data=f.read(),
                    file_name="tonghop_khoa_capnhat.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            st.success(f"Đã truyền dữ liệu hoạt động quy đổi từ file GV sang file tổng hợp Khoa cho giáo viên '{selected_gv}'.")
        else:
            st.error(f"Không tìm thấy tên giáo viên '{selected_gv}' trong cột B của sheet CAC_HOAT_DONG.")
    import os
    os.unlink(gv_path)
    os.unlink(khoa_path)
