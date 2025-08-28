# thoi_khoa_bieu.py

import streamlit as st
import pandas as pd
import openpyxl
import io
import re
import gspread
from google.oauth2.service_account import Credentials
from unidecode import unidecode
from thefuzz import process

# --- CÁC HẰNG SỐ ---
# XÓA: Hằng số này không còn cần thiết vì ta sẽ tải trực tiếp từ Google Sheet
# COMMON_SUBJECTS = [
#     "Giáo dục chính trị", "Pháp luật", "Giáo dục thể chất",
#     "Giáo dục quốc phòng và an ninh", "Tin học", "Tiếng Anh"
# ]
# SIMILARITY_THRESHOLD = 85

# --- CÁC HÀM KẾT NỐI GOOGLE SHEETS ---

@st.cache_resource
def connect_to_gsheet():
    """Kết nối tới Google Sheets sử dụng service account credentials."""
    try:
        creds_dict = st.secrets["gcp_service_account"]
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive.file"
        ]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        return gspread.authorize(creds)
    except Exception as e:
        st.error(f"Lỗi kết nối Google Sheets: {e}")
        return None

@st.cache_data(ttl=600)
def load_abbreviations_map(_gsheet_client, spreadsheet_id):
    """Tải bản đồ ánh xạ viết tắt từ sheet VIET_TAT."""
    if _gsheet_client is None: return {}
    try:
        spreadsheet = _gsheet_client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet("VIET_TAT")
        records = worksheet.get_all_records()
        abbreviations_map = {
            str(record.get('Viết_tắt_1')).strip().lower(): str(record.get('Đầy_đủ_1')).strip()
            for record in records if record.get('Viết_tắt_1') and record.get('Đầy_đủ_1')
        }
        return abbreviations_map
    except gspread.exceptions.WorksheetNotFound:
        st.warning("⚠️ Không tìm thấy sheet 'VIET_TAT'. Tính năng chuẩn hóa viết tắt sẽ bị bỏ qua.")
        return {}
    except Exception as e:
        st.warning(f"⚠️ Lỗi khi tải danh sách viết tắt từ sheet 'VIET_TAT': {e}")
        return {}

# MỚI: Hàm tải danh sách các môn chung từ sheet DANH_MUC
@st.cache_data(ttl=600)
def load_common_subjects_map(_gsheet_client, spreadsheet_id):
    """Tải bản đồ ánh xạ Môn học chung và Mã môn chung từ sheet DANH_MUC."""
    if _gsheet_client is None: return {}
    try:
        spreadsheet = _gsheet_client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet("DANH_MUC")
        records = worksheet.get_all_records()
        # Chuẩn hóa tên môn học về chữ thường để dễ so sánh
        common_subjects_map = {
            str(record.get('Tên_mônchung')).strip().lower(): str(record.get('Mã_mônchung')).strip()
            for record in records if record.get('Tên_mônchung') and record.get('Mã_mônchung')
        }
        return common_subjects_map
    except gspread.exceptions.WorksheetNotFound:
        st.warning("⚠️ Không tìm thấy sheet 'DANH_MUC'. Không thể tự động gán mã môn chung.")
        return {}
    except Exception as e:
        st.warning(f"⚠️ Lỗi khi tải danh sách môn chung từ sheet 'DANH_MUC': {e}")
        return {}

@st.cache_data(ttl=600)
def load_teacher_info(_gsheet_client, spreadsheet_id):
    """Tải và chuẩn hóa dữ liệu giáo viên từ sheet THONG_TIN_GV."""
    if _gsheet_client is None: return pd.DataFrame()
    try:
        spreadsheet = _gsheet_client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet("THONG_TIN_GV")
        df = pd.DataFrame(worksheet.get_all_records())
        df.columns = df.columns.str.strip()
        required_cols = ['Ho_ten_gv', 'Khoa', 'Ma_gv']
        missing_cols = [col for col in required_cols if col not in df.columns]

        if missing_cols:
            st.error(
                f"**Lỗi Cấu Trúc File: Thiếu các cột bắt buộc trong sheet 'THONG_TIN_GV'.**\n\n"
                f"Các cột bị thiếu là: **`{', '.join(missing_cols)}`**.\n\n"
                f"*Vui lòng kiểm tra lại tên cột trong file Google Sheet của bạn.*\n\n"
                f"Các cột hiện tại tìm thấy là: `{', '.join(df.columns)}`"
            )
            return pd.DataFrame()

        df['Ho_ten_gv_normalized'] = df['Ho_ten_gv'].astype(str).apply(lambda x: unidecode(x).lower())
        df['First_name'] = df['Ho_ten_gv'].astype(str).apply(lambda x: x.split(' ')[-1])
        df['First_name_normalized'] = df['First_name'].astype(str).apply(lambda x: unidecode(x).lower())
        return df
    except Exception as e:
        st.error(f"Lỗi khi tải dữ liệu giáo viên: {e}")
        return pd.DataFrame()

@st.cache_data(ttl=60)
def get_khoa_list(_gsheet_client, spreadsheet_id):
    if _gsheet_client is None: return []
    try:
        spreadsheet = _gsheet_client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet("DANH_MUC")
        df = pd.DataFrame(worksheet.get_all_records())
        return df["Khoa/Phòng/Trung tâm"].dropna().unique().tolist()
    except Exception as e:
        st.error(f"Lỗi khi tải danh sách khoa: {e}")
        return []

def update_gsheet_by_khoa(client, spreadsheet_id, sheet_name, df_new, khoa_to_update):
    try:
        spreadsheet = client.open_by_key(spreadsheet_id)
        try:
            worksheet = spreadsheet.worksheet(sheet_name)
            existing_data = worksheet.get_all_records()
            if existing_data:
                df_existing = pd.DataFrame(existing_data)
                df_others = df_existing[df_existing['KHOA'] != khoa_to_update]
                df_combined = pd.concat([df_others, df_new], ignore_index=True)
            else:
                df_combined = df_new
        except gspread.WorksheetNotFound:
            df_combined = df_new
            worksheet = spreadsheet.add_worksheet(title=sheet_name, rows="1", cols="1")

        worksheet.clear()
        data_to_upload = [df_combined.columns.values.tolist()] + df_combined.astype(str).values.tolist()
        worksheet.update(data_to_upload, 'A1')
        return True, None
    except Exception as e:
        return False, str(e)

def bulk_update_teacher_info(gsheet_client, spreadsheet_id, updates_list):
    if not updates_list: return True, "Không có tên viết tắt mới cần cập nhật."
    try:
        spreadsheet = gsheet_client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet("THONG_TIN_GV")
        headers = worksheet.row_values(1)
        try:
            col_index = headers.index('Ten_viet_tat') + 1
        except ValueError:
            return False, "Không tìm thấy cột 'Ten_viet_tat' trong sheet THONG_TIN_GV."
        cell_updates = [gspread.Cell(row=update['index'] + 2, col=col_index, value=update['value']) for update in updates_list]
        if cell_updates:
            worksheet.update_cells(cell_updates, value_input_option='USER_ENTERED')
            return True, f"Đã cập nhật thành công {len(cell_updates)} tên viết tắt mới."
        return True, "Không có tên viết tắt mới cần cập nhật."
    except Exception as e:
        return False, f"Lỗi khi cập nhật hàng loạt tên viết tắt: {e}"

# --- CÁC HÀM XỬ LÝ EXCEL ---

def extract_schedule_from_excel(worksheet):
    ngay_ap_dung = ""
    for r_idx in range(1, 6):
        for c_idx in range(1, 27):
            cell_value = str(worksheet.cell(row=r_idx, column=c_idx).value or '').strip()
            if "áp dụng" in cell_value.lower():
                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', cell_value)
                if date_match:
                    ngay_ap_dung = date_match.group(1)
                else:
                    try:
                        next_cell_value = str(worksheet.cell(row=r_idx, column=c_idx + 1).value or '')
                        if re.search(r'(\d{1,2}/\d{1,2}/\d{4})', next_cell_value):
                            ngay_ap_dung = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', next_cell_value).group(1)
                    except: pass
                break
        if ngay_ap_dung: break

    start_row, start_col = -1, -1
    for r_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10), 1):
        for c_idx, cell in enumerate(row, 1):
            if cell.value and "thứ" in str(cell.value).lower():
                start_row, start_col = r_idx, c_idx; break
        if start_row != -1: break
    if start_row == -1:
        st.error(f"Không tìm thấy ô tiêu đề 'Thứ' trong sheet '{worksheet.title}'."); return None, None

    last_row = start_row
    for r_idx in range(worksheet.max_row, start_row - 1, -1):
        if worksheet.cell(row=r_idx, column=start_col + 2).value is not None:
            last_row = r_idx; break
    last_col = start_col
    for row in worksheet.iter_rows(min_row=start_row, max_row=last_row):
        for cell in row:
            if cell.value is not None and cell.column > last_col: last_col = cell.column

    merged_values = {}
    for merged_range in worksheet.merged_cells.ranges:
        top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
        for row_ in range(merged_range.min_row, merged_range.max_row + 1):
            for col_ in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row_, col_)] = top_left_cell.value

    day_to_number_map = {'HAI': 2, 'BA': 3, 'TƯ': 4, 'NĂM': 5, 'SÁU': 6, 'BẢY': 7}
    data = []
    for r_idx in range(start_row, last_row + 1):
        row_data = [merged_values.get((r_idx, c_idx), worksheet.cell(row=r_idx, column=c_idx).value) for c_idx in range(start_col, last_col + 1)]
        if r_idx > start_row:
            clean_day = re.sub(r'\s+', '', str(row_data[0] or '')).strip().upper()
            row_data[0] = day_to_number_map.get(clean_day, row_data[0])
        data.append(row_data)

    if not data: return None, ngay_ap_dung

    header_level1, header_level2 = data[0], data[1]
    filled_header_level1 = []
    last_val = ""
    for val in header_level1:
        if val is not None and str(val).strip() != '': last_val = val
        filled_header_level1.append(last_val)

    combined_headers = [f"{str(h1 or '').strip()}___{str(h2 or '').strip()}" if i >= 3 else str(h1 or '').strip() for i, (h1, h2) in enumerate(zip(filled_header_level1, header_level2))]
    df = pd.DataFrame(data[2:], columns=combined_headers)
    return df, ngay_ap_dung

# --- HÀM LOGIC CHÍNH ---

# XÓA: Toàn bộ hàm này sẽ được tích hợp vào `transform_to_database_format`
# def normalize_common_subjects(df, abbreviations_map):
# ...

def create_teacher_mapping(df_schedule, df_teacher_info_full, selected_khoa):
    """Tạo bản đồ ánh xạ tên GV và danh sách cần cập nhật."""
    def get_all_individual_names(series):
        names = set()
        for item in series.dropna():
            for name in str(item).split(' / '):
                if name.strip():
                    names.add(name.strip())
        return names

    all_bm_names = get_all_individual_names(df_schedule['Giáo viên BM'])
    all_cn_names = get_all_individual_names(df_schedule['Giáo viên CN'])
    all_short_names = all_bm_names.union(all_cn_names)

    df_teachers_in_khoa = df_teacher_info_full[df_teacher_info_full['Khoa'] == selected_khoa].copy()
    mapping, updates_for_gsheet = {}, []

    if 'Ten_viet_tat' in df_teachers_in_khoa.columns:
        df_with_shortnames = df_teachers_in_khoa.dropna(subset=['Ten_viet_tat'])
        df_with_shortnames = df_with_shortnames[df_with_shortnames['Ten_viet_tat'].astype(str).str.strip() != '']
        for _, row in df_with_shortnames.iterrows():
            for short_name in str(row['Ten_viet_tat']).split(';'):
                sn_clean = short_name.strip()
                if sn_clean and sn_clean not in mapping:
                    mapping[sn_clean] = {'full_name': row['Ho_ten_gv'], 'id': row['Ma_gv']}

    for short_name in all_short_names:
        if short_name in mapping: continue
        match = re.match(r'([TC])\.(.*)', short_name)
        if not match:
            mapping[short_name] = {'full_name': short_name, 'id': ''}; continue

        prefix, name_part = match.groups()
        name_part_normalized = unidecode(name_part.strip()).lower()
        possible_matches = df_teachers_in_khoa[df_teachers_in_khoa['First_name_normalized'] == name_part_normalized]

        if len(possible_matches) == 1:
            matched_teacher = possible_matches.iloc[0]
            mapping[short_name] = {'full_name': matched_teacher['Ho_ten_gv'], 'id': matched_teacher['Ma_gv']}
            original_ten_viet_tat = df_teacher_info_full.loc[matched_teacher.name, 'Ten_viet_tat']
            if pd.isna(original_ten_viet_tat) or str(original_ten_viet_tat).strip() == '':
                updates_for_gsheet.append({'index': matched_teacher.name, 'value': short_name})
        else:
            mapping[short_name] = {'full_name': short_name, 'id': ''}
    return mapping, updates_for_gsheet

# SỬA ĐỔI: Hàm được tái cấu trúc hoàn toàn để tích hợp logic chuẩn hóa
def transform_to_database_format(df_wide, ngay_ap_dung, abbreviations_map, common_subjects_map):
    id_vars = ['Thứ', 'Buổi', 'Tiết']
    df_long = pd.melt(df_wide, id_vars=id_vars, var_name='Lớp_Raw', value_name='Chi tiết Môn học')
    df_long.dropna(subset=['Chi tiết Môn học'], inplace=True)
    df_long = df_long[df_long['Chi tiết Môn học'].astype(str).str.strip() != '']

    def parse_and_normalize_subject(cell_text):
        """
        Hàm lồng nhau để bóc tách, chuẩn hóa và tra cứu mã môn từ một ô.
        Trả về: (Tên môn đầy đủ, Mã môn, Phòng học, Tên GV, Ghi chú)
        """
        clean_text = re.sub(r'\s{2,}', ' ', str(cell_text).replace('\n', ' ').strip())
        ghi_chu, remaining_text = "", clean_text

        # 1. Tách Ghi chú (Học từ... / Chỉ học...)
        note_match = re.search(r'\(?((?:Học từ|Chỉ học).*?)\)?$', clean_text, re.IGNORECASE)
        if note_match:
            ghi_chu = note_match.group(1).strip()
            remaining_text = clean_text.replace(note_match.group(0), '').strip()
        
        # Xử lý trường hợp đặc biệt
        if "THPT" in remaining_text.upper():
            return ("HỌC TKB VĂN HÓA THPT", "", "", "", ghi_chu)

        entries = [e.strip() for e in remaining_text.split(';')]
        all_mon_hoc, all_ma_mon, all_phong_hoc, all_gv = [], [], [], []

        for entry in entries:
            if not entry: continue
            
            mon_hoc_raw, phong_hoc, gv_part = entry, "", ""
            
            # 2. Tách Môn học, Phòng học, Giáo viên
            match = re.search(r'^(.*?)\s*\((.*)\)$', entry)
            if match:
                mon_hoc_raw, content_in_parens = match.group(1).strip(), match.group(2).strip()
                if '-' in content_in_parens:
                    parts = content_in_parens.split('-', 1)
                    phong_hoc, gv_part = parts[0].strip(), parts[1].strip()
                else:
                    phong_hoc, gv_part = "", content_in_parens
            
            # 3. Chuẩn hóa tên môn học (ưu tiên VIET_TAT)
            mon_hoc_normalized_key = mon_hoc_raw.lower()
            mon_hoc_day_du = abbreviations_map.get(mon_hoc_normalized_key, mon_hoc_raw)

            # 4. Tra cứu Mã môn chung từ DANH_MUC
            ma_mon = common_subjects_map.get(mon_hoc_day_du.lower(), '')

            all_mon_hoc.append(mon_hoc_day_du)
            if ma_mon: all_ma_mon.append(ma_mon)
            if phong_hoc: all_phong_hoc.append(phong_hoc)
            if gv_part:
                gv_list = [g.strip() for g in gv_part.split('/') if g.strip()]
                all_gv.extend(gv_list)

        # Tổng hợp kết quả từ các entry trong một ô
        final_mon_hoc = "; ".join(sorted(list(set(all_mon_hoc))))
        final_ma_mon = " / ".join(sorted(list(set(all_ma_mon))))
        final_phong_hoc = " / ".join(sorted(list(set(all_phong_hoc))))
        final_gv = " / ".join(sorted(list(set(all_gv))))
        
        return (final_mon_hoc, final_ma_mon, final_phong_hoc, final_gv, ghi_chu)


    parsed_cols = df_long['Chi tiết Môn học'].apply(parse_and_normalize_subject)
    mh_extracted = pd.DataFrame(parsed_cols.tolist(), index=df_long.index, columns=['Môn học', 'Mã môn', 'Phòng học', 'Giáo viên BM', 'Ghi chú'])
    header_parts = df_long['Lớp_Raw'].str.split('___', expand=True)
    lop_extracted = header_parts[0].str.extract(r'^(.*?)\s*(?:\((\d+)\))?$'); lop_extracted.columns = ['Lớp', 'Sĩ số']

    def parse_cn_details(text):
        if not text or pd.isna(text): return ("", "", "")
        text = str(text).replace('(', '').replace(')', '')
        parts = text.split('-')
        return (parts[0].strip(), parts[1].strip() if len(parts) > 1 else "", "")

    cn_details = header_parts[1].apply(parse_cn_details) if len(header_parts.columns) > 1 else pd.Series([("", "", "")] * len(df_long))
    cn_extracted = pd.DataFrame(cn_details.tolist(), index=df_long.index, columns=['Phòng SHCN', 'Giáo viên CN', 'Lớp VHPT'])
    df_final = pd.concat([df_long[['Thứ', 'Buổi', 'Tiết']].reset_index(drop=True), lop_extracted.reset_index(drop=True), cn_extracted.reset_index(drop=True), mh_extracted.reset_index(drop=True)], axis=1)
    df_final['Trình độ'] = df_final['Lớp'].apply(lambda x: 'Cao đẳng' if 'C.' in str(x) else ('Trung Cấp' if 'T.' in str(x) else ''))
    df_final.fillna('', inplace=True)
    
    # Các cột mã GV và Khoa sẽ được điền ở các bước sau
    df_final['Ma_gv_bm'] = ''
    df_final['Ma_gv_cn'] = ''
    df_final['KHOA'] = ''
    df_final['Ngày áp dụng'] = ngay_ap_dung
    
    final_cols = ['Thứ', 'Buổi', 'Tiết', 'Lớp', 'Sĩ số', 'Trình độ', 
                  'Môn học', 'Mã môn', 'Phòng học', 
                  'Giáo viên BM', 'Ma_gv_bm', 
                  'Phòng SHCN', 'Giáo viên CN', 'Ma_gv_cn', 
                  'Lớp VHPT', 'Ghi chú', 'KHOA', 'Ngày áp dụng']
    return df_final[final_cols]

# --- Giao diện chính của ứng dụng Streamlit ---

st.set_page_config(page_title="Quản lý TKB", layout="wide")
st.title("📥 Tải lên & Xử lý Thời Khóa Biểu")

if 'unmapped_teachers' not in st.session_state: st.session_state.unmapped_teachers = None
if 'final_df_to_save' not in st.session_state: st.session_state.final_df_to_save = None

TEACHER_INFO_SHEET_ID = "1TJfaywQM1VNGjDbWyC3osTLLOvlgzP0-bQjz8J-_BoI"
gsheet_client = None
abbreviations_map = {}
common_subjects_map = {} # MỚI

if "gcp_service_account" in st.secrets:
    gsheet_client = connect_to_gsheet()
    if gsheet_client:
        if 'df_teacher_info' not in st.session_state:
            st.session_state.df_teacher_info = load_teacher_info(gsheet_client, TEACHER_INFO_SHEET_ID)
        abbreviations_map = load_abbreviations_map(gsheet_client, TEACHER_INFO_SHEET_ID)
        common_subjects_map = load_common_subjects_map(gsheet_client, TEACHER_INFO_SHEET_ID) # MỚI
else:
    st.warning("Không tìm thấy cấu hình Google Sheets trong `st.secrets`.", icon="⚠️")

uploaded_file = st.file_uploader("Chọn file Excel TKB của bạn", type=["xlsx"])

if uploaded_file:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(uploaded_file.getvalue()), data_only=True)
        all_sheet_names = workbook.sheetnames
        sheets_to_display = [s for s in all_sheet_names if s.upper() not in ["DANH_MUC", "THONG_TIN_GV", "VIET_TAT"]]
        selected_sheets = st.multiselect("Chọn các sheet TKB cần xử lý:", options=sheets_to_display)

        if st.button("Xử lý các sheet đã chọn", key="process_button"):
            st.session_state.unmapped_teachers = None
            st.session_state.final_df_to_save = None
            all_processed_dfs = []
            ngay_ap_dung_dict = {}

            with st.spinner("Đang xử lý dữ liệu..."):
                for sheet_name in selected_sheets:
                    worksheet = workbook[sheet_name]
                    raw_df, ngay_ap_dung = extract_schedule_from_excel(worksheet)
                    if raw_df is not None:
                        if ngay_ap_dung: ngay_ap_dung_dict[sheet_name] = ngay_ap_dung
                        # SỬA ĐỔI: Truyền các map cần thiết vào hàm
                        db_df = transform_to_database_format(raw_df, ngay_ap_dung, abbreviations_map, common_subjects_map)
                        all_processed_dfs.append(db_df)

            if all_processed_dfs:
                combined_df = pd.concat(all_processed_dfs, ignore_index=True)
                # XÓA: Bước chuẩn hóa riêng biệt này không còn cần thiết
                # with st.spinner("Đang rà soát và chuẩn hóa các môn học chung..."):
                #     st.session_state['processed_df'] = normalize_common_subjects(combined_df, abbreviations_map)
                st.session_state['processed_df'] = combined_df # SỬA ĐỔI: Gán trực tiếp dataframe đã xử lý

                st.success("Xử lý file Excel thành công!")
                st.info("Đã tự động bóc tách, chuẩn hóa tên môn và tra cứu mã môn chung.")
                if ngay_ap_dung_dict:
                    st.write("Đã tìm thấy ngày áp dụng trong các sheet sau:")
                    for sheet, date in ngay_ap_dung_dict.items():
                        st.info(f"- Sheet **'{sheet}'**: {date}")
            else:
                st.warning("Không thể trích xuất dữ liệu từ các sheet đã chọn.")

# Phần còn lại của code giữ nguyên...
        if 'processed_df' in st.session_state:
            db_df_to_process = st.session_state['processed_df']
            st.markdown("---")
            st.subheader("📤 Lưu trữ dữ liệu đã xử lý")
            st.info(f"Dữ liệu sẽ được lưu vào Google Sheet có ID: **{TEACHER_INFO_SHEET_ID}**")

            col1, col2, col3, col4 = st.columns(4)
            with col1: nam_hoc = st.text_input("Năm học:", value="2425", key="nh")
            with col2: hoc_ky = st.text_input("Học kỳ:", value="HK1", key="hk")
            with col3: giai_doan = st.text_input("Giai đoạn:", value="GD1", key="gd")
            with col4:
                khoa_list = get_khoa_list(gsheet_client, TEACHER_INFO_SHEET_ID)
                khoa = st.selectbox("Khoa:", options=khoa_list, key="khoa")

            sheet_name = f"DATA_{nam_hoc}_{hoc_ky}_{giai_doan}"
            st.write(f"Tên sheet sẽ được tạo/cập nhật là: **{sheet_name}**")

            if st.button("1. Bắt đầu ánh xạ và kiểm tra", key="start_mapping_button"):
                if gsheet_client and khoa and not db_df_to_process.empty and not st.session_state.df_teacher_info.empty:
                    with st.spinner(f"Đang tự động ánh xạ GV cho khoa '{khoa}'..."):
                        teacher_mapping, updates_list = create_teacher_mapping(db_df_to_process, st.session_state.df_teacher_info, khoa)
                        
                        df_after_mapping = db_df_to_process.copy()
                        df_after_mapping['KHOA'] = khoa

                        def apply_mapping(name_str, key):
                            name_str = str(name_str).strip()
                            if not name_str: return ""
                            names_list = [n.strip() for n in name_str.split(' / ')]
                            mapped_names = [str(teacher_mapping.get(name, {}).get(key, name)) for name in names_list]
                            return " / ".join(mapped_names)

                        df_after_mapping['Ma_gv_bm'] = df_after_mapping['Giáo viên BM'].apply(apply_mapping, key='id')
                        df_after_mapping['Giáo viên BM'] = df_after_mapping['Giáo viên BM'].apply(apply_mapping, key='full_name')
                        df_after_mapping['Ma_gv_cn'] = df_after_mapping['Giáo viên CN'].apply(apply_mapping, key='id')
                        df_after_mapping['Giáo viên CN'] = df_after_mapping['Giáo viên CN'].apply(apply_mapping, key='full_name')
                        
                        unmapped_bm = df_after_mapping[df_after_mapping['Giáo viên BM'].str.contains(r'^[TC]\.', na=False)][['Giáo viên BM', 'Môn học']]
                        unmapped_cn = df_after_mapping[df_after_mapping['Giáo viên CN'].str.contains(r'^[TC]\.', na=False)][['Giáo viên CN', 'Lớp']]
                        
                        unmapped_teachers = {}
                        for _, row in unmapped_bm.iterrows():
                            for name in row['Giáo viên BM'].split(' / '):
                                if re.match(r'^[TC]\.', name.strip()):
                                    unmapped_teachers[name.strip()] = f"Môn: {row['Môn học']}"
                        for _, row in unmapped_cn.iterrows():
                            for name in row['Giáo viên CN'].split(' / '):
                                if re.match(r'^[TC]\.', name.strip()):
                                    unmapped_teachers[name.strip()] = f"CN Lớp: {row['Lớp']}"
                        
                        st.session_state.final_df_to_save = df_after_mapping
                        st.session_state.updates_list = updates_list

                        if unmapped_teachers:
                            st.warning(f"Phát hiện {len(unmapped_teachers)} giáo viên không thể tự động ánh xạ. Vui lòng xử lý thủ công.")
                            st.session_state.unmapped_teachers = unmapped_teachers
                        else:
                            st.success("Tuyệt vời! Tất cả giáo viên đã được tự động ánh xạ thành công.")
                            st.session_state.unmapped_teachers = {}
                else:
                    st.error("Không thể bắt đầu. Vui lòng chọn một Khoa và đảm bảo dữ liệu giáo viên đã được tải thành công.")

            if st.session_state.unmapped_teachers is not None:
                unmapped_teachers = st.session_state.unmapped_teachers
                df_teachers_in_khoa = st.session_state.df_teacher_info[st.session_state.df_teacher_info['Khoa'] == khoa]
                teacher_list_in_khoa = ["-- Chọn --"] + df_teachers_in_khoa['Ho_ten_gv'].tolist()

                if not unmapped_teachers:
                     if st.button("2. Hoàn tất và Lưu vào Google Sheet", key="final_save_no_manual"):
                        with st.spinner("Đang lưu dữ liệu..."):
                            if st.session_state.updates_list:
                                success_update, msg_update = bulk_update_teacher_info(gsheet_client, TEACHER_INFO_SHEET_ID, st.session_state.updates_list)
                                if success_update: st.info(msg_update)
                                else: st.error(msg_update)
                            success_save, msg_save = update_gsheet_by_khoa(gsheet_client, TEACHER_INFO_SHEET_ID, sheet_name, st.session_state.final_df_to_save, khoa)
                            if success_save:
                                st.success(f"Cập nhật dữ liệu TKB vào sheet '{sheet_name}' thành công!")
                                st.cache_data.clear()
                                st.session_state.unmapped_teachers = None
                            else: st.error(f"Lỗi khi lưu TKB: {msg_save}")
                else:
                    st.markdown("---")
                    st.subheader("✍️ Xử lý thủ công các giáo viên chưa được ánh xạ")
                    with st.form("manual_update_form"):
                        manual_selections = {}
                        for short_name, context in unmapped_teachers.items():
                            cols = st.columns([1, 2, 3])
                            with cols[0]: st.write(f"**{short_name}**")
                            with cols[1]: st.caption(context)
                            with cols[2]: manual_selections[short_name] = st.selectbox(f"Chọn tên đúng cho {short_name}", options=teacher_list_in_khoa, label_visibility="collapsed", key=f"select_{short_name}")
                        
                        if st.form_submit_button("2. Xác nhận lựa chọn và Lưu vào Google Sheet"):
                            with st.spinner("Đang cập nhật lựa chọn và lưu dữ liệu..."):
                                df_to_save = st.session_state.final_df_to_save.copy()
                                final_updates = {item['index']: item['value'] for item in st.session_state.get('updates_list', [])}
                                for short_name, full_name in manual_selections.items():
                                    if full_name != "-- Chọn --":
                                        teacher_info = df_teachers_in_khoa[df_teachers_in_khoa['Ho_ten_gv'] == full_name].iloc[0]
                                        teacher_id = str(teacher_info['Ma_gv'])
                                        
                                        df_to_save['Giáo viên BM'] = df_to_save['Giáo viên BM'].str.replace(short_name, full_name, regex=False)
                                        df_to_save['Ma_gv_bm'] = df_to_save['Ma_gv_bm'].str.replace(short_name, teacher_id, regex=False)
                                        df_to_save['Giáo viên CN'] = df_to_save['Giáo viên CN'].str.replace(short_name, full_name, regex=False)
                                        df_to_save['Ma_gv_cn'] = df_to_save['Ma_gv_cn'].str.replace(short_name, teacher_id, regex=False)
                                        
                                        current_short_names_str = str(st.session_state.df_teacher_info.loc[teacher_info.name, 'Ten_viet_tat'])
                                        if short_name not in [s.strip() for s in current_short_names_str.split(';')]:
                                            new_val = short_name if (pd.isna(current_short_names_str) or not current_short_names_str.strip()) else f"{current_short_names_str};{short_name}"
                                            final_updates[teacher_info.name] = new_val
                                
                                final_updates_list = [{'index': k, 'value': v} for k, v in final_updates.items()]
                                if final_updates_list:
                                    success_update, msg_update = bulk_update_teacher_info(gsheet_client, TEACHER_INFO_SHEET_ID, final_updates_list)
                                    if success_update: st.info(msg_update)
                                    else: st.error(msg_update)
                                
                                success_save, msg_save = update_gsheet_by_khoa(gsheet_client, TEACHER_INFO_SHEET_ID, sheet_name, df_to_save, khoa)
                                if success_save:
                                    st.success(f"Cập nhật dữ liệu TKB vào sheet '{sheet_name}' thành công!")
                                    st.cache_data.clear()
                                    st.session_state.unmapped_teachers = None
                                else: st.error(f"Lỗi khi lưu TKB: {msg_save}")
            
            with st.expander("Xem trước dữ liệu đã xử lý (trước khi ánh xạ GV)"):
                st.dataframe(db_df_to_process)

    except Exception as e:
        st.error(f"Đã có lỗi xảy ra khi xử lý file: {e}")
        st.exception(e)
