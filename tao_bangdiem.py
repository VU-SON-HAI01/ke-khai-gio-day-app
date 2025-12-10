import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
# Sửa lỗi chính tả 'impoart' -> 'import'
from openpyxl.styles import Border, Side, Font
from openpyxl import load_workbook
import io
import re
import zipfile

# --- CÁC HÀM HỖ TRỢ ---

def find_student_data_in_sheet(worksheet):
    """
    Tìm và trích xuất dữ liệu học sinh từ một sheet có cấu trúc không cố định.
    - Tự động tìm dòng header dựa vào 'STT' ở cột A hoặc B.
    - Sau đó tìm cột 'Họ và tên' và 'Năm sinh' trên dòng header đó.
    - Chuẩn hóa và tách riêng 2 cột họ và tên.
    - Chuẩn hóa và định dạng cột ngày sinh.
    - Dừng lại khi CẢ HAI cột họ và tên đều trống hoặc chứa số.
    - Trả về một DataFrame.
    """
    header_row_index = -1
    name_col_index = -1
    dob_col_index = -1
    student_data = []
    found_end_row = False

    # 1. Tìm dòng header
    for i, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        col_a_val = str(row[0]).lower().strip() if len(row) > 0 and row[0] is not None else ''
        col_b_val = str(row[1]).lower().strip() if len(row) > 1 and row[1] is not None else ''
        if 'stt' in col_a_val or 'stt' in col_b_val:
            header_row_index = i
            header_content = [str(cell).lower().strip() if cell is not None else '' for cell in row]
            break
    if header_row_index == -1:
        st.warning(f"Không thể tìm thấy dòng tiêu đề (header) chứa 'STT' trong sheet '{worksheet.title}'.")
        return None

    # 2. Xác định vị trí các cột cần thiết
    ten_dem_col_index = None
    for idx, col in enumerate(header_content):
        if col in ["họ đệm", "họ dem"]:
            ten_dem_col_index = idx + 1
            break
    if ten_dem_col_index is None:
        for idx, col in enumerate(header_content):
            if col in ["họ và tên", "ho va ten"]:
                ten_dem_col_index = idx + 1
                break
    if ten_dem_col_index is None:
        st.error(f"Trong sheet '{worksheet.title}', đã tìm thấy dòng tiêu đề ở dòng {header_row_index} nhưng thiếu cột bắt buộc. Lỗi: không tìm thấy cột 'Họ đệm' hoặc 'Họ và Tên'.")
        return None
    ten_col_index = ten_dem_col_index + 1
    dob_col_index = None
    for idx, col in enumerate(header_content):
        if col in ["năm sinh", "nam sinh"]:
            dob_col_index = idx + 1
            break
    if dob_col_index is None:
        for idx, col in enumerate(header_content):
            if col in ["ngày sinh", "ngay sinh"]:
                dob_col_index = idx + 1
                break
    if dob_col_index is None:
        st.error(f"Trong sheet '{worksheet.title}', đã tìm thấy dòng tiêu đề ở dòng {header_row_index} nhưng thiếu cột bắt buộc. Lỗi: không tìm thấy cột 'Năm sinh' hoặc 'Ngày sinh'.")
        return None

    # 3. Đọc dữ liệu
    # Dừng lại nếu 2 dòng liên tiếp tiếp theo (cột 'TÊN') đều rỗng/None hoặc là số, hoặc 1 dòng là số và dòng sau là rỗng
    rows = list(worksheet.iter_rows(min_row=header_row_index + 1, values_only=True))
    i = 0
    while i < len(rows):
        row = rows[i]
        ten_dem_cell = row[ten_dem_col_index - 1]
        ten_cell = row[ten_col_index - 1]
        dob_cell = row[dob_col_index - 1]

        # Kiểm tra điều kiện dừng: 2 dòng tiếp theo cột 'TÊN' đều rỗng/None/number hoặc 1 dòng là số, dòng sau là rỗng
        stop = False
        if i + 1 < len(rows):
            next_row1 = rows[i]
            next_row2 = rows[i + 1]
            ten1 = next_row1[ten_col_index - 1]
            ten2 = next_row2[ten_col_index - 1]
            ten1_empty_or_number = (ten1 is None or str(ten1).strip() == '' or isinstance(ten1, (int, float)))
            ten2_empty_or_number = (ten2 is None or str(ten2).strip() == '' or isinstance(ten2, (int, float)))
            # Dừng nếu cả hai đều rỗng/None/number
            if ten1_empty_or_number and ten2_empty_or_number:
                found_end_row = True
                break
            # Dừng nếu dòng 1 là số, dòng 2 là rỗng
            if (isinstance(ten1, (int, float)) and (ten2 is None or str(ten2).strip() == '')):
                found_end_row = True
                break

        # Nếu dòng hiện tại là rỗng/None/number thì bỏ qua
        if ten_cell is None or str(ten_cell).strip() == '' or isinstance(ten_cell, (int, float)):
            i += 1
            continue

        ten_dem_str = re.sub(r'\s+', ' ', str(ten_dem_cell or '')).strip()
        ten_str = re.sub(r'\s+', ' ', str(ten_cell or '')).strip()

        formatted_dob = ''
        if dob_cell is not None:
            try:
                dt_object = pd.to_datetime(dob_cell, errors='coerce')
                if pd.notna(dt_object):
                    formatted_dob = dt_object.strftime('%d/%m/%Y')
                else:
                    formatted_dob = str(dob_cell).strip()
            except Exception:
                formatted_dob = str(dob_cell).strip()

        student_data.append({
            "TÊN ĐỆM": ten_dem_str,
            "TÊN": ten_str,
            "NGÀY SINH": formatted_dob
        })
        i += 1

    # Chỉ cảnh báo nếu không có dữ liệu học sinh nào được trích xuất
    if not found_end_row and len(student_data) == 0:
        st.warning(f"Không tìm thấy dữ liệu học sinh hợp lệ trong sheet '{worksheet.title}'.")

    return pd.DataFrame(student_data)


def check_data_consistency(data_file, danh_muc_file):
    """
    Kiểm tra sự khớp nhau giữa các sheet trong file dữ liệu và danh mục lớp.
    """
    try:
        xls_data = pd.ExcelFile(data_file)
        data_sheet_names = set(xls_data.sheet_names)

        xls_danh_muc = pd.ExcelFile(danh_muc_file)
        if "DANH_MUC" not in xls_danh_muc.sheet_names:
            st.error("File Danh mục thiếu sheet 'DANH_MUC'.")
            return None, None
        
        df_danh_muc = pd.read_excel(xls_danh_muc, sheet_name="DANH_MUC")
        valid_class_names = set(df_danh_muc.iloc[:, 1].dropna().astype(str))

        sheets_not_in_danh_muc = data_sheet_names - valid_class_names
        danh_muc_not_in_sheets = valid_class_names - data_sheet_names

        return sheets_not_in_danh_muc, danh_muc_not_in_sheets
    except Exception as e:
        st.error(f"Lỗi khi kiểm tra dữ liệu: {e}")
        return None, None


def process_excel_files(template_file, data_file, danh_muc_file, hoc_ky, nam_hoc, cap_nhat):
    """
    Hàm chính để xử lý, chèn dữ liệu từ file data vào file template.
    """
    generated_files = {}
    skipped_sheets = []
    
    # --- Tải dữ liệu từ file Danh mục (Cải tiến để chống lỗi) ---
    try:
        xls_danh_muc = pd.ExcelFile(danh_muc_file)
        
        if "DANH_MUC" not in xls_danh_muc.sheet_names:
            st.error(f"Lỗi: Không tìm thấy sheet 'DANH_MUC' trong file DS LOP(Mau).xlsx. Các sheet có sẵn: {xls_danh_muc.sheet_names}")
            return {}, []
        
        if "DATA_GOC" not in xls_danh_muc.sheet_names:
            st.error(f"Lỗi: Không tìm thấy sheet 'DATA_GOC' trong file DS LOP(Mau).xlsx. Các sheet có sẵn: {xls_danh_muc.sheet_names}")
            return {}, []
            
        df_danh_muc = pd.read_excel(xls_danh_muc, sheet_name="DANH_MUC")
        df_data_goc = pd.read_excel(xls_danh_muc, sheet_name="DATA_GOC", header=1)
        
        # Lấy danh sách các lớp hợp lệ từ cột B
        valid_class_names = set(df_danh_muc.iloc[:, 1].dropna().astype(str))

    except Exception as e:
        st.error(f"Lỗi khi đọc File Danh mục Lớp (DS LOP(Mau).xlsx): {e}")
        return {}, []
        
    data_workbook = openpyxl.load_workbook(data_file, data_only=True)
    
    for sheet_name in data_workbook.sheetnames:
        # *** KIỂM TRA TÍNH HỢP LỆ CỦA SHEET ***
        if sheet_name not in valid_class_names:
            skipped_sheets.append(sheet_name)
            continue # Bỏ qua sheet này và chuyển sang sheet tiếp theo

        worksheet = data_workbook[sheet_name]

        df_sheet_data = find_student_data_in_sheet(worksheet)
        
        if df_sheet_data is None or df_sheet_data.empty:
            st.warning(f"Không tìm thấy dữ liệu học sinh hợp lệ trong sheet '{sheet_name}'. Bỏ qua sheet này.")
            continue

        class_info = df_danh_muc[df_danh_muc.iloc[:, 1] == sheet_name]
        # Không cần kiểm tra class_info.empty nữa vì đã kiểm tra ở trên
        
        nganh_nghe = class_info.iloc[0, 3]
        ma_nghe = str(class_info.iloc[0, 4])

        template_file.seek(0)
        output_workbook = openpyxl.load_workbook(template_file)
        
        # --- XỬ LÝ SHEET "Bang diem qua trinh" ---
        try:
            output_sheet_qt = output_workbook["Bang diem qua trinh"]
            output_sheet_qt.protection.set_password('PDT')
        except KeyError:
            st.error("Lỗi: File mẫu không chứa sheet có tên 'Bang diem qua trinh'.")
            return {}, skipped_sheets

        try:
            hoc_ky_numeric = int(hoc_ky)
        except (ValueError, TypeError):
            hoc_ky_numeric = hoc_ky

        output_sheet_qt.cell(row=2, column=9).value = sheet_name
        output_sheet_qt.cell(row=3, column=9).value = hoc_ky_numeric
        output_sheet_qt.cell(row=4, column=9).value = nam_hoc
        output_sheet_qt.cell(row=3, column=28).value = cap_nhat
        output_sheet_qt.cell(row=2, column=20).value = nganh_nghe

        list_mon_hoc = []
        target_col_name = None
        for col in df_data_goc.columns:
            if ma_nghe in str(col):
                target_col_name = col
                break
        
        if target_col_name:
            list_mon_hoc = df_data_goc[target_col_name].dropna().astype(str).tolist()
        else:
            st.warning(f"Không tìm thấy cột môn học cho mã nghề '{ma_nghe}' trong sheet DATA_GOC.")

        if list_mon_hoc:
            dv_sheet_name = "DSMON"
            try:
                dv_sheet = output_workbook[dv_sheet_name]
                if dv_sheet.max_row > 1:
                    dv_sheet.delete_rows(idx=2, amount=dv_sheet.max_row - 1)
            except KeyError:
                st.warning(f"File mẫu không có sheet '{dv_sheet_name}'. Sẽ tạo một sheet mới.")
                dv_sheet = output_workbook.create_sheet(dv_sheet_name)
                dv_sheet.cell(row=1, column=1).value = "STT"
                dv_sheet.cell(row=1, column=2).value = "DSMON"

            for i, mon_hoc in enumerate(list_mon_hoc, 1):
                row_index = i + 1
                dv_sheet.cell(row=row_index, column=1).value = i
                dv_sheet.cell(row=row_index, column=2).value = mon_hoc
                
            formula = f"'{dv_sheet_name}'!$B$2:$B${len(list_mon_hoc) + 1}" 
            
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.error = 'Giá trị không hợp lệ.'
            dv.errorTitle = 'Dữ liệu không hợp lệ'
            dv.prompt = 'Vui lòng chọn từ danh sách'
            dv.promptTitle = 'Chọn Môn học'
            output_sheet_qt.add_data_validation(dv)
            dv.add('V1')
            dv_sheet.sheet_state = 'hidden'

        num_students = len(df_sheet_data)
        EXTRA_BLANK_ROWS = 2 
        total_rows_needed = num_students + EXTRA_BLANK_ROWS
        
        QT_START_ROW = 7
        QT_TEMPLATE_STUDENT_ROWS = 5
        QT_INSERT_BEFORE_ROW = 12
        QT_STYLE_ROW = 9
        QT_BORDER_END_COL = 30

        rows_to_insert_qt = total_rows_needed - QT_TEMPLATE_STUDENT_ROWS
        if rows_to_insert_qt > 0:
            output_sheet_qt.insert_rows(QT_INSERT_BEFORE_ROW, amount=rows_to_insert_qt)
            for row_idx in range(QT_INSERT_BEFORE_ROW, QT_INSERT_BEFORE_ROW + rows_to_insert_qt):
                for col_idx in range(1, output_sheet_qt.max_column + 1):
                    source_cell = output_sheet_qt.cell(row=QT_STYLE_ROW, column=col_idx)
                    new_cell = output_sheet_qt.cell(row=row_idx, column=col_idx)
                    if source_cell.has_style:
                        new_cell.font = source_cell.font.copy()
                        new_cell.border = source_cell.border.copy()
                        new_cell.fill = source_cell.fill.copy()
                        new_cell.number_format = source_cell.number_format
                        new_cell.protection = source_cell.protection.copy()
                        new_cell.alignment = source_cell.alignment.copy()

        formulas_qt = {}
        for col in range(16, output_sheet_qt.max_column + 1):
            cell = output_sheet_qt.cell(row=QT_START_ROW, column=col)
            if cell.value and str(cell.value).startswith('='):
                formulas_qt[col] = cell.value
        for row_num in range(QT_START_ROW, QT_START_ROW + total_rows_needed):
            for col_num, formula_str in formulas_qt.items():
                new_formula = formula_str.replace(str(QT_START_ROW), str(row_num))
                output_sheet_qt.cell(row=row_num, column=col_num).value = new_formula

        for i, student_row in df_sheet_data.iterrows():
            current_row_index = QT_START_ROW + i
            output_sheet_qt.cell(row=current_row_index, column=1).value = i + 1
            output_sheet_qt.cell(row=current_row_index, column=3).value = student_row["TÊN ĐỆM"]
            output_sheet_qt.cell(row=current_row_index, column=4).value = student_row["TÊN"]
            output_sheet_qt.cell(row=current_row_index, column=5).value = student_row["NGÀY SINH"]

        last_data_row_qt = QT_START_ROW + total_rows_needed - 1
        double_line_side = Side(style='double')
        for col_idx in range(1, QT_BORDER_END_COL + 1):
            cell_to_border = output_sheet_qt.cell(row=last_data_row_qt, column=col_idx)
            existing_border = cell_to_border.border
            cell_to_border.border = Border(left=existing_border.left, right=existing_border.right, top=existing_border.top, bottom=double_line_side)

        # --- XỬ LÝ SHEET "Bang diem thi" ---
        try:
            output_sheet_thi = output_workbook["Bang diem thi"]
            output_sheet_thi.protection.set_password('PDT')
            
            THI_DATA_START_ROW = 10
            THI_TEMPLATE_ROW = 11
            THI_TEMPLATE_STUDENT_ROWS = 5
            THI_INSERT_BEFORE_ROW = 15
            THI_FILL_END_COL = 25
            
            rows_to_insert_thi = total_rows_needed - THI_TEMPLATE_STUDENT_ROWS
            if rows_to_insert_thi > 0:
                output_sheet_thi.insert_rows(THI_INSERT_BEFORE_ROW, amount=rows_to_insert_thi)
            
            template_styles = {}
            template_formulas = {}
            for col_idx in range(1, THI_FILL_END_COL + 1):
                template_cell = output_sheet_thi.cell(row=THI_TEMPLATE_ROW, column=col_idx)
                if template_cell.has_style:
                    template_styles[col_idx] = template_cell
                if template_cell.value and str(template_cell.value).startswith('='):
                    template_formulas[col_idx] = template_cell.value
            
            for row_num in range(THI_DATA_START_ROW, THI_DATA_START_ROW + total_rows_needed):
                row_offset = row_num - THI_TEMPLATE_ROW
                for col_idx in range(1, THI_FILL_END_COL + 1):
                    target_cell = output_sheet_thi.cell(row=row_num, column=col_idx)

                    if col_idx in template_styles:
                        source_cell_for_style = template_styles[col_idx]
                        target_cell.font = source_cell_for_style.font.copy()
                        target_cell.border = source_cell_for_style.border.copy()
                        target_cell.fill = source_cell_for_style.fill.copy()
                        target_cell.number_format = source_cell_for_style.number_format
                        target_cell.protection = source_cell_for_style.protection.copy()
                        target_cell.alignment = source_cell_for_style.alignment.copy()

                    if col_idx in template_formulas:
                        formula_str = template_formulas[col_idx]
                        
                        def adjust_row_reference(match):
                            col_part = match.group(1)
                            row_abs = match.group(2)
                            row_num_str = match.group(3)
                            if row_abs: return match.group(0)
                            else:
                                new_row = int(row_num_str) + row_offset
                                return f"{col_part}{new_row}"

                        pattern = re.compile(r"(\$?[A-Z]{1,3})(\$?)(\d+)")
                        new_formula = pattern.sub(adjust_row_reference, formula_str)
                        target_cell.value = new_formula
            
            last_data_row_thi = THI_DATA_START_ROW + total_rows_needed - 1
            for col_idx in range(1, THI_FILL_END_COL + 1):
                cell_to_border = output_sheet_thi.cell(row=last_data_row_thi, column=col_idx)
                existing_border = cell_to_border.border
                cell_to_border.border = Border(left=existing_border.left, right=existing_border.right, top=existing_border.top, bottom=double_line_side)

        except KeyError:
            st.warning("File mẫu không chứa sheet 'Bang diem thi'. Bỏ qua xử lý sheet này.")

        # *** TẠO TÊN FILE MỚI ***
        clean_cap_nhat = cap_nhat.replace('-', '_')
        final_file_name = f"{sheet_name}_bangdiem_{clean_cap_nhat}.xlsx"
        
        output_buffer = io.BytesIO()
        output_workbook.save(output_buffer)
        generated_files[final_file_name] = output_buffer.getvalue()
        
    return generated_files, skipped_sheets

# --- GIAO DIỆN ỨNG DỤNG STREAMLIT ---

st.title("⚙️ Công cụ Cập nhật Bảng điểm HSSV")
st.markdown("---")

# Khởi tạo session state
if 'generated_files' not in st.session_state:
    st.session_state.generated_files = {}
if 'skipped_sheets' not in st.session_state:
    st.session_state.skipped_sheets = []
if 'zip_buffer' not in st.session_state:
    st.session_state.zip_buffer = None


import datetime

st.header("Thông tin chung")
col1, col2, col3 = st.columns(3)

# Tính giá trị mặc định cho năm học và cập nhật
now = datetime.datetime.now()
current_year = now.year
current_month = now.month
if current_month >= 8 and current_month <= 12:
    nam_hoc_default = f"{current_year}-{current_year+1}"
else:
    nam_hoc_default = f"{current_year-1}-{current_year}"
cap_nhat_default = f"T{current_month}-{current_year}"

with col1:
    hoc_ky_input = st.text_input("Học kỳ", value="1")
with col2:
    nam_hoc_input = st.text_input("Năm học", value=nam_hoc_default)
with col3:
    cap_nhat_input = st.text_input("Cập nhật", value=cap_nhat_default)
st.markdown("---")

with st.expander("Tải lên (file mẫu) khác với file mẫu mặc định"):
    import os
    sample_path = "data_base/Bang_diem_qua_trinh_(Mau).xlsx"
    if os.path.exists(sample_path):
        with open(sample_path, "rb") as f:
            st.download_button(
                label="📥 Tải xuống Mẫu bảng điểm",
                data=f.read(),
                file_name="Bang_diem_qua_trinh_(Mau).xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning(f"Không tìm thấy file mẫu: {sample_path}")
    uploaded_template_file = st.file_uploader(
        "1. 📂 Tải lên File Mẫu Bảng Điểm (.xlsx)",
        type=['xlsx'],
        key="template_uploader"
    )
    sample_danhmuc_path = "data_base/DS_LOP_(Mau).xlsx"
    if os.path.exists(sample_danhmuc_path):
        with open(sample_danhmuc_path, "rb") as f:
            st.download_button(
                label="📥 Tải xuống Mẫu danh mục lớp",
                data=f.read(),
                file_name="DS_LOP_(Mau).xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning(f"Không tìm thấy file mẫu: {sample_danhmuc_path}")
    uploaded_danh_muc_file = st.file_uploader(
        "2. 📂 Tải lên File Danh mục Lớp (DS LOP(Mau).xlsx)",
        type=['xlsx'],
        key="danh_muc_uploader"
    )

with st.container():
    st.subheader("Bước 1: Tải dữ liệu danh sách sinh viên", divider=True)

    uploaded_data_file = st.file_uploader(
    "1. 📂 Tải lên File Dữ Liệu HSSV (.xlsx)",
    type=['xlsx'],
    key="data_uploader"
    )
    
    # Thêm selector chọn Khóa học
    st.markdown("---")
    khoa_options = ["K49", "K50", "K51"]
    selected_khoa = st.selectbox("Chọn Khóa học để xử lý", khoa_options, key="khoa_selector")

    # Lọc dữ liệu theo khóa học đã chọn
    def is_class_in_khoa(class_name, khoa):
        return class_name.startswith(khoa[1:])

    filtered_data = []
    if 'all_data' not in st.session_state:
        st.session_state.all_data = []
    if st.session_state.all_data:
        for df in st.session_state.all_data:
            if "Tên lớp" in df.columns:
                df_khoa = df[df["Tên lớp"].apply(lambda x: is_class_in_khoa(str(x), selected_khoa))]
                if not df_khoa.empty:
                    filtered_data.append(df_khoa)
        if filtered_data:
            df_filtered = pd.concat(filtered_data, ignore_index=True)
            st.session_state.filtered_data = filtered_data
            st.session_state.df_filtered = df_filtered
            st.subheader(f"Danh sách lớp thuộc {selected_khoa}")
            st.dataframe(df_filtered, use_container_width=True)
        else:
            df_filtered = pd.DataFrame()
            st.session_state.filtered_data = []
            st.session_state.df_filtered = df_filtered
            st.info(f"Không có lớp nào thuộc {selected_khoa} trong dữ liệu đã xử lý.")
    else:
        df_filtered = pd.DataFrame()
        st.session_state.filtered_data = []
        st.session_state.df_filtered = df_filtered
    st.subheader("Bước 2: Kiểm tra & Xử lý", divider=True)
    # Container để hiển thị kết quả kiểm tra
    check_results_placeholder = st.container()

    if uploaded_data_file:
        if st.button("🔍 Kiểm tra dữ liệu", use_container_width=True, key="btn_kiem_tra_du_lieu_main"):
            # Nếu chưa upload danh mục thì dùng file mặc định
            danh_muc_file_obj = uploaded_danh_muc_file
            if danh_muc_file_obj is None:
                danh_muc_file_obj = open("data_base/DS_LOP_(Mau).xlsx", "rb")
            # Lọc chỉ các sheet thuộc khóa đã chọn
            xls_data = pd.ExcelFile(uploaded_data_file)
            all_sheet_names = xls_data.sheet_names
            khoa_prefix = selected_khoa[1:]  # VD: 'K51' -> '51'
            sheet_names_to_check = [name for name in all_sheet_names if str(name).startswith(khoa_prefix)]
            # Đọc danh mục
            xls_danh_muc = pd.ExcelFile(danh_muc_file_obj)
            df_danh_muc = pd.read_excel(xls_danh_muc, sheet_name="DANH_MUC")
            # Chỉ lấy các lớp trong danh mục thuộc khóa đã chọn
            valid_class_names = set(df_danh_muc.iloc[:, 1].dropna().astype(str))
            valid_class_names_khoa = set([name for name in valid_class_names if str(name).startswith(khoa_prefix)])
            sheets_not_in_danh_muc = set(sheet_names_to_check) - valid_class_names_khoa
            danh_muc_not_in_sheets = valid_class_names_khoa - set(sheet_names_to_check)
            if uploaded_danh_muc_file is None:
                danh_muc_file_obj.close()
            with check_results_placeholder:
                if not sheets_not_in_danh_muc and not danh_muc_not_in_sheets:
                    st.success("✅ Dữ liệu hợp lệ! Tất cả các sheet đều khớp với danh mục.")
                if sheets_not_in_danh_muc:
                    st.warning("⚠️ Các sheet sau có trong file dữ liệu nhưng không có trong danh mục và sẽ bị bỏ qua:")
                    st.json(list(sheets_not_in_danh_muc))
                if danh_muc_not_in_sheets:
                    st.info("ℹ️ Các lớp sau có trong danh mục nhưng không có sheet tương ứng trong file dữ liệu:")
                    st.json(list(danh_muc_not_in_sheets))
    
    if uploaded_data_file:
        if st.button("🚀 Xử lý và Tạo Files", type="primary", use_container_width=True, key="btn_xuly_tao_files_main"):
            st.session_state.zip_buffer = None
            try:
                # Nếu không upload file mẫu thì dùng file mẫu mặc định
                template_file_obj = uploaded_template_file
                if not df_filtered.empty:
                        if st.button("🔍 Kiểm tra dữ liệu", use_container_width=True, key="btn_kiem_tra_du_lieu_inner"):
                        # Nếu chưa upload danh mục thì dùng file mặc định
                            danh_muc_file_obj = uploaded_danh_muc_file
                            if danh_muc_file_obj is None:
                                danh_muc_file_obj = open("data_base/DS_LOP_(Mau).xlsx", "rb")
                            # Chỉ kiểm tra các lớp thuộc khóa đã chọn
                            sheet_names_to_check = set(df_filtered["Tên lớp"].unique())
                            xls_danh_muc = pd.ExcelFile(danh_muc_file_obj)
                            df_danh_muc = pd.read_excel(xls_danh_muc, sheet_name="DANH_MUC")
                            valid_class_names = set(df_danh_muc.iloc[:, 1].dropna().astype(str))
                            sheets_not_in_danh_muc = sheet_names_to_check - valid_class_names
                            danh_muc_not_in_sheets = valid_class_names - sheet_names_to_check
                            if uploaded_danh_muc_file is None:
                                danh_muc_file_obj.close()
                            with check_results_placeholder:
                                if not sheets_not_in_danh_muc and not danh_muc_not_in_sheets:
                                    st.success("✅ Dữ liệu hợp lệ! Tất cả các sheet đều khớp với danh mục.")
                                if sheets_not_in_danh_muc:
                                    st.warning("⚠️ Các sheet sau có trong dữ liệu nhưng không có trong danh mục và sẽ bị bỏ qua:")
                                    st.json(list(sheets_not_in_danh_muc))
                                if danh_muc_not_in_sheets:
                                    st.info("ℹ️ Các lớp sau có trong danh mục nhưng không có sheet tương ứng trong dữ liệu:")
                                    st.json(list(danh_muc_not_in_sheets))

                if uploaded_data_file is not None:
                    # Lọc chỉ các sheet thuộc khóa đã chọn
                    xls_data = pd.ExcelFile(uploaded_data_file)
                    all_sheet_names = xls_data.sheet_names
                    khoa_prefix = selected_khoa[1:]  # VD: 'K51' -> '51'
                    sheet_names_to_process = [name for name in all_sheet_names if str(name).startswith(khoa_prefix)]
                    # Đọc từng sheet, gom dữ liệu
                    all_data = []
                    for sheet in sheet_names_to_process:
                        try:
                            df = pd.read_excel(uploaded_data_file, sheet_name=sheet)
                            df["Tên lớp"] = sheet
                            all_data.append(df)
                        except Exception:
                            continue
                    if all_data:
                        df_filtered = pd.concat(all_data, ignore_index=True)
                        st.session_state.df_filtered = df_filtered
                        st.session_state.all_data = all_data
                    else:
                        df_filtered = pd.DataFrame()
                        st.session_state.df_filtered = df_filtered
                        st.session_state.all_data = []
                    if not df_filtered.empty:
                        if st.button("🚀 Xử lý và Tạo Files", type="primary", use_container_width=True, key="btn_xuly_tao_files_inner"):
                            st.session_state.zip_buffer = None
                            try:
                                template_file_obj = uploaded_template_file
                                if template_file_obj is None:
                                    template_file_obj = open("data_base/Bang_diem_qua_trinh_(Mau).xlsx", "rb")
                                danh_muc_file_obj = uploaded_danh_muc_file
                                if danh_muc_file_obj is None:
                                    danh_muc_file_obj = open("data_base/DS_LOP_(Mau).xlsx", "rb")
                                st.session_state.generated_files, st.session_state.skipped_sheets = process_excel_files(
                                    template_file_obj,
                                    uploaded_data_file,
                                    danh_muc_file_obj,
                                    hoc_ky_input,
                                    nam_hoc_input,
                                    cap_nhat_input
                                )
                                if uploaded_template_file is None:
                                    template_file_obj.close()
                                if uploaded_danh_muc_file is None:
                                    danh_muc_file_obj.close()
                                if st.session_state.generated_files:
                                    st.success(f"✅ Hoàn thành! Đã xử lý và tạo ra {len(st.session_state.generated_files)} file.")
                                    with st.spinner("Đang nén file..."):
                                        zip_buffer = io.BytesIO()
                                        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED, False) as zf:
                                            for file_name, file_data in st.session_state.generated_files.items():
                                                zf.writestr(file_name, file_data)
                                        st.session_state.zip_buffer = zip_buffer
                                else:
                                    st.warning("Quá trình xử lý hoàn tất nhưng không có file nào được tạo. Vui lòng kiểm tra lại các file đầu vào.")
                                if st.session_state.skipped_sheets:
                                    st.info(f"ℹ️ Các sheet sau đã bị bỏ qua vì không có trong danh mục: {', '.join(st.session_state.skipped_sheets)}")
                            except Exception as e:
                                st.error(f"Đã xảy ra lỗi trong quá trình xử lý: {e}")
                    else:
                        st.info("Không có dữ liệu lớp nào để gom.")
            except Exception as e:
                st.error(f"Không đọc được dữ liệu lớp. Chi tiết lỗi: {e}")
            from openpyxl.utils.dataframe import dataframe_to_rows
            # Chỉ gom dữ liệu đã được lọc theo khóa
            df_filtered = st.session_state.df_filtered if 'df_filtered' in st.session_state else pd.DataFrame()
            st.dataframe(df_filtered, use_container_width=True)
            mau_path = "data_base/mau_thong_tin_nguoi_hoc.xlsx"
            if os.path.exists(mau_path):
                if st.button("Gom dữ liệu", use_container_width=True, key="btn_gom_du_lieu"):
                    # Gom dữ liệu học sinh đã được xử lý/chuẩn hóa từ từng sheet (dùng find_student_data_in_sheet)
                    if uploaded_data_file is not None:
                        xls_data = pd.ExcelFile(uploaded_data_file)
                        all_sheet_names = xls_data.sheet_names
                        khoa_prefix = selected_khoa[1:]
                        sheet_names_to_process = [name for name in all_sheet_names if str(name).startswith(khoa_prefix)]
                        # Đọc danh mục để lọc lớp hợp lệ
                        danh_muc_file_obj = uploaded_danh_muc_file
                        if danh_muc_file_obj is None:
                            danh_muc_file_obj = open("data_base/DS_LOP_(Mau).xlsx", "rb")
                        xls_danh_muc = pd.ExcelFile(danh_muc_file_obj)
                        df_danh_muc = pd.read_excel(xls_danh_muc, sheet_name="DANH_MUC")
                        valid_class_names = set(df_danh_muc.iloc[:, 1].dropna().astype(str))
                        valid_class_names_khoa = set([name for name in valid_class_names if str(name).startswith(khoa_prefix)])
                        # Chỉ lấy các sheet hợp lệ
                        sheet_names_to_process = [name for name in sheet_names_to_process if name in valid_class_names_khoa]
                        # Đọc từng sheet, gom dữ liệu học sinh đã chuẩn hóa
                        import openpyxl
                        wb_data = openpyxl.load_workbook(uploaded_data_file, data_only=True)
                        all_student_rows = []
                        for sheet in sheet_names_to_process:
                            ws = wb_data[sheet]
                            df_students = find_student_data_in_sheet(ws)
                            if df_students is not None and not df_students.empty:
                                df_students = df_students.copy()
                                df_students["Tên lớp"] = sheet
                                all_student_rows.append(df_students)
                        if all_student_rows:
                            df_all_students = pd.concat(all_student_rows, ignore_index=True)
                            wb = load_workbook(mau_path)
                            ws = wb.active
                            ws.delete_rows(2, ws.max_row - 1)
                            for r in dataframe_to_rows(df_all_students, index=False, header=False):
                                ws.append(r)
                            output = io.BytesIO()
                            wb.save(output)
                            st.session_state.updated_mau_file = output
                            st.success("Đã gom dữ liệu học sinh đã chuẩn hóa vào file mẫu!")
                        else:
                            st.warning("Không có dữ liệu học sinh hợp lệ nào để gom.")
                        if uploaded_danh_muc_file is None:
                            danh_muc_file_obj.close()
                    else:
                        st.warning("Không có file dữ liệu lớp nào để gom.")
            if st.session_state.get("updated_mau_file"):
                st.download_button(
                    label="Tải về file mau_thong_tin_nguoi_hoc.xlsx đã gom",
                    data=st.session_state.updated_mau_file.getvalue(),
                    file_name="mau_thong_tin_nguoi_hoc.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
    else:
        st.info("Chưa có file nào được tạo để gom dữ liệu.")

