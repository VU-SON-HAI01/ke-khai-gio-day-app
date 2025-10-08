
import streamlit as st
import pandas as pd
import openpyxl
import pandas as pd
import os

# Hàm tổng hợp kết quả từ các trang kê khai
# Giả định dữ liệu đã được lưu vào session_state từ các trang kê khai


def export_giangday_to_excel(spreadsheet=None, df_mon=None, df_hk1=None, template_path='data_base/mau_kegio.xlsx'):
    """
    Kết hợp: Nếu truyền spreadsheet và df_mon thì lấy dữ liệu từ Google Sheet, ánh xạ tên môn học và ghi vào file Excel mẫu.
    Nếu truyền df_hk1 thì ghi trực tiếp dữ liệu HK1 vào file Excel mẫu.
    """

    if not os.path.exists(template_path):
        return False, f'Không tìm thấy file mẫu: {template_path}. Hãy upload file mau_kegio.xlsx vào thư mục data_base.'
    wb = openpyxl.load_workbook(template_path)
    if 'Ke_gio_HK1' in wb.sheetnames:
        sheet = wb['Ke_gio_HK1']
    else:
        sheet = wb.active
    start_row = 8
    # Nếu truyền spreadsheet và df_mon: lấy dữ liệu từ Google Sheet
    if spreadsheet is not None and df_mon is not None:
        ws = next((ws for ws in spreadsheet.worksheets() if ws.title == 'output_giangday'), None)
        if ws is None:
            return False, "Không tìm thấy sheet 'output_giangday' trong Google Sheet."
        df = pd.DataFrame(ws.get_all_records())
        for i, row in df.iterrows():
            excel_row = int(start_row + i)
            # Đảm bảo excel_row không vượt quá số dòng tối đa của sheet
            if excel_row < 1:
                continue
            # Ghi dữ liệu vào từng cell
            try:
                sheet.cell(row=excel_row, column=1).value = row.get('Tuần', '')      # A
                # Ghi giá trị cột 'Lớp_học' vào từng dòng (cột 2)
                sheet.cell(row=excel_row, column=2).value = row.get('Lớp_học', '')  # B
                sheet.cell(row=excel_row, column=3).value = row.get('Sĩ số', '')    # C
                ma_mon_nganh = row.get('Mã_Môn_Ngành', '')
                mon_hoc = ''
                if not df_mon.empty and 'Mã_môn_ngành' in df_mon.columns:
                    mon_row = df_mon[df_mon['Mã_môn_ngành'] == ma_mon_nganh]
                    if not mon_row.empty and 'Môn_học' in mon_row.columns:
                        mon_hoc = mon_row.iloc[0]['Môn_học']
                sheet.cell(row=excel_row, column=4).value = mon_hoc                 # D
                sheet.cell(row=excel_row, column=5).value = row.get('HS TC/CĐ', '') # E
                sheet.cell(row=excel_row, column=6).value = row.get('Tiết', '')     # F
                sheet.cell(row=excel_row, column=7).value = row.get('Tiết_LT', '')  # G
                sheet.cell(row=excel_row, column=8).value = row.get('Tiết_TH', '')  # H
                sheet.cell(row=excel_row, column=9).value = row.get('HS_SS_LT', '') # I
                sheet.cell(row=excel_row, column=10).value = row.get('HS_SS_TH', '')# J
                # Cột K: Thực hành nghề Nặng nhọc
                nang_nhoc_val = ''
                if not df_mon.empty and 'Mã_môn_ngành' in df_mon.columns and 'Nặng_nhọc' in df_mon.columns:
                    ma_mon_nganh = row.get('Mã_Môn_Ngành', '')
                    mon_row = df_mon[df_mon['Mã_môn_ngành'] == ma_mon_nganh]
                    if not mon_row.empty:
                        nn_val = mon_row.iloc[0]['Nặng_nhọc']
                        if nn_val == 'NN':
                            nang_nhoc_val = 'NN'
                sheet.cell(row=excel_row, column=11).value = nang_nhoc_val  # K
            except Exception as e:
                print(f"Lỗi ghi dòng {excel_row}: {e}")
                continue
    # Nếu truyền df_hk1: ghi trực tiếp dữ liệu HK1
    elif df_hk1 is not None:
        for i, row in df_hk1.iterrows():
            excel_row = int(start_row + i)
            if excel_row < 1:
                continue
            try:
                sheet.cell(row=excel_row, column=1).value = row.get('Tuần', '')      # A
                sheet.cell(row=excel_row, column=3).value = row.get('Sĩ số', '')    # C
                sheet.cell(row=excel_row, column=4).value = row.get('Môn_học', '')  # D
                sheet.cell(row=excel_row, column=5).value = row.get('HS TC/CĐ', '') # E
                sheet.cell(row=excel_row, column=6).value = row.get('Tiết', '')     # F
                sheet.cell(row=excel_row, column=7).value = row.get('Tiết_LT', '')  # G
                sheet.cell(row=excel_row, column=8).value = row.get('Tiết_TH', '')  # H
                sheet.cell(row=excel_row, column=9).value = row.get('HS_SS_LT', '') # I
                sheet.cell(row=excel_row, column=10).value = row.get('HS_SS_TH', '')# J
            except Exception as e:
                print(f"Lỗi ghi dòng {excel_row}: {e}")
                continue
    wb.save(template_path)
    return True, template_path
def tonghop_ketqua():
    # Nút tải dữ liệu từ Google Sheet của user (các sheet có tên bắt đầu bằng 'output_')
    if 'export_ready' not in st.session_state:
        st.session_state['export_ready'] = False
    dfs = []
    load_clicked = st.button("Xem kết quả dư giờ", use_container_width=True)

    if load_clicked:
        spreadsheet = st.session_state.get('spreadsheet')
        if spreadsheet is None:
            st.error("Không tìm thấy file Google Sheet của bạn trong session_state. Hãy đăng nhập lại hoặc liên hệ Admin.")
            return
        try:
            sheet_list = spreadsheet.worksheets()
            # Định nghĩa thứ tự và tên hiển thị
            sheet_order = [
                ("output_giangday", "✍️ Bảng tổng hợp khối lượng dạy"),
                ("output_thiketthuc", "📝 Bảng tổng hợp khối thi kết thúc"),
                ("output_quydoigiam", "⚖️ Bảng tổng hợp Giảm trừ/Kiêm nhiệm"),
                ("output_hoatdong", "🏃 Bảng tổng hợp Kê Hoạt động quy đổi khác")
            ]
            found_any = False
            for idx, (sheet_name, display_name) in enumerate(sheet_order):
                ws = next((ws for ws in sheet_list if ws.title == sheet_name), None)
                if ws is not None:
                    df_raw = ws.get_all_records()
                    df = pd.DataFrame(df_raw)
                    if not df.empty:
                        # Chỉ ẩn bảng output_giangday, các bảng khác vẫn hiển thị
                        if sheet_name != "output_giangday":
                            st.subheader(display_name)
                            # Ẩn các cột theo yêu cầu từng bảng
                            df_display = df.copy()
                            if sheet_name == "output_thiketthuc":
                                for col in ["Mã HĐ", "Mã NCKH"]:
                                    if col in df_display.columns:
                                        df_display = df_display.drop(columns=[col])
                            elif sheet_name == "output_quydoigiam":
                                for col in ["Mã HĐ", "Mã NCKH", "activity_index"]:
                                    if col in df_display.columns:
                                        df_display = df_display.drop(columns=[col])
                                # Thêm dòng Tổng cộng vào cuối bảng
                                if "Giờ quy đổi" in df_display.columns:
                                    tongcong = df_display["Giờ quy đổi"].apply(pd.to_numeric, errors='coerce').sum()
                                    total_row = {col: '' for col in df_display.columns}
                                    if "Nội dung hoạt động" in df_display.columns:
                                        total_row["Nội dung hoạt động"] = "Tổng cộng"
                                    total_row["Giờ quy đổi"] = tongcong
                                    df_display = pd.concat([df_display, pd.DataFrame([total_row])], ignore_index=True)
                            elif sheet_name == "output_hoatdong":
                                # Ẩn cả "Mã NCKH" và "MÃ NCKH" nếu có
                                for col in ["Mã HĐ", "Mã NCKH", "MÃ NCKH", "activity_index"]:
                                    if col in df_display.columns:
                                        df_display = df_display.drop(columns=[col])
                                # Thêm dòng Tổng cộng vào cuối bảng
                                if "Giờ quy đổi" in df_display.columns:
                                    tongcong = df_display["Giờ quy đổi"].apply(pd.to_numeric, errors='coerce').sum()
                                    total_row = {col: '' for col in df_display.columns}
                                    if "Hoạt động quy đổi" in df_display.columns:
                                        total_row["Hoạt động quy đổi"] = "Tổng cộng"
                                    total_row["Giờ quy đổi"] = tongcong
                                    df_display = pd.concat([df_display, pd.DataFrame([total_row])], ignore_index=True)
                            st.dataframe(df_display)
                        # Nếu là bảng giảng dạy, chỉ tạo bảng tổng hợp HK1/HK2 mà không hiển thị bảng gốc
                        if sheet_name == "output_giangday":
                            import numpy as np
                            df_gd = df.copy()
                            input_gd = None
                            try:
                                input_gd_ws = next((ws for ws in sheet_list if ws.title == 'input_giangday'), None)
                                if input_gd_ws is not None:
                                    input_gd = pd.DataFrame(input_gd_ws.get_all_records())
                            except Exception:
                                input_gd = None
                            if 'ID_MÔN' not in df_gd.columns:
                                st.warning('Không tìm thấy cột ID_MÔN trong dữ liệu output_giangday.')
                            else:
                                mon_list = df_gd['ID_MÔN'].unique()
                                rows = []
                                for mon in mon_list:
                                    df_mon = df_gd[df_gd['ID_MÔN'] == mon]
                                    if df_mon.empty:
                                        continue
                                    lop_mon = ''
                                    if input_gd is not None and 'ID_MÔN' in input_gd.columns:
                                        row_map = input_gd[input_gd['ID_MÔN'] == mon]
                                        if not row_map.empty:
                                            lop = row_map.iloc[0]['lop_hoc'] if 'lop_hoc' in row_map.columns else ''
                                            mon_name = row_map.iloc[0]['mon_hoc'] if 'mon_hoc' in row_map.columns else ''
                                            lop_mon = f"{lop} // {mon_name}"
                                    if not lop_mon:
                                        lop_mon = mon
                                    tuan_min = df_mon['Tuần'].iloc[0] if 'Tuần' in df_mon.columns else ''
                                    tuan_max = df_mon['Tuần'].iloc[-1] if 'Tuần' in df_mon.columns else ''
                                    tuan_str = f"T{tuan_min} - T{tuan_max}" if tuan_min != '' and tuan_max != '' else ''
                                    si_so = df_mon['Sĩ số'].iloc[-1] if 'Sĩ số' in df_mon.columns else ''
                                    tiet = df_mon['Tiết'].sum() if 'Tiết' in df_mon.columns else 0.0
                                    tiet_lt = df_mon['Tiết_LT'].sum() if 'Tiết_LT' in df_mon.columns else 0.0
                                    tiet_th = df_mon['Tiết_TH'].sum() if 'Tiết_TH' in df_mon.columns else 0.0
                                    qd_thua = df_mon['QĐ thừa'].sum() if 'QĐ thừa' in df_mon.columns else 0.0
                                    qd_thieu = df_mon['QĐ thiếu'].sum() if 'QĐ thiếu' in df_mon.columns else 0.0
                                    try:
                                        tuan_min_num = float(tuan_min)
                                        tuan_max_num = float(tuan_max)
                                        avg_tuan = (tuan_min_num + tuan_max_num) / 2
                                        hoc_ky = 2 if avg_tuan > 22 else 1
                                    except Exception:
                                        hoc_ky = 1
                                    rows.append({
                                        'Lớp // Môn': lop_mon,
                                        'Tuần': tuan_str,
                                        'Sĩ số': si_so,
                                        'Tiết': tiet,
                                        'Tiết LT': tiet_lt,
                                        'Tiết TH': tiet_th,
                                        'QĐ thừa': qd_thua,
                                        'QĐ Thiếu': qd_thieu,
                                        'Học kỳ': hoc_ky
                                    })
                                df_tonghop_mon = pd.DataFrame(rows)
                                st.subheader('✍️ Bảng tổng hợp khối lượng dạy')
                                for hk in [1, 2]:
                                    df_hk = df_tonghop_mon[df_tonghop_mon['Học kỳ'] == hk].copy()
                                    if not df_hk.empty:
                                        for col in ['Tiết', 'Tiết LT', 'Tiết TH', 'QĐ thừa', 'QĐ Thiếu']:
                                            df_hk[col] = pd.to_numeric(df_hk[col], errors='coerce').fillna(0.0)
                                        total_row = {
                                            'Lớp // Môn': 'Tổng cộng',
                                            'Tuần': '',
                                            'Sĩ số': '',
                                            'Tiết': df_hk['Tiết'].sum(),
                                            'Tiết LT': df_hk['Tiết LT'].sum(),
                                            'Tiết TH': df_hk['Tiết TH'].sum(),
                                            'QĐ thừa': df_hk['QĐ thừa'].sum(),
                                            'QĐ Thiếu': df_hk['QĐ Thiếu'].sum(),
                                            'Học kỳ': ''
                                        }
                                        df_hk = pd.concat([df_hk, pd.DataFrame([total_row])], ignore_index=True)
                                        st.markdown(f"**Bảng tổng hợp tiết giảng dạy quy đổi HK{hk}**")
                                        st.dataframe(df_hk.drop(columns=['Học kỳ']), use_container_width=True)
                                        if hk == 1:
                                            st.session_state['df_hk1'] = df_hk
                                        elif hk == 2:
                                            st.session_state['df_hk2'] = df_hk
                    dfs.append(df)
                    found_any = True
            if dfs:
                st.subheader(":blue[BẢNG TỔNG HỢP KHỐI LƯỢNG DƯ/THIẾU GIỜ]")
                giochuan = st.session_state.get('giochuan', 616)
                def build_bang_tonghop(dfs, giochuan=616):
                    import numpy as np
                    tiet_giangday_hk1_qdthieu = 0
                    tiet_giangday_hk1_qdthua = 0
                    tiet_giangday_hk2_qdthieu = 0
                    tiet_giangday_hk2_qdthua = 0
                    df_hk1 = st.session_state.get('df_hk1')
                    df_hk2 = st.session_state.get('df_hk2')
                    if df_hk1 is not None and not df_hk1.empty:
                        row_total = df_hk1[df_hk1['Lớp // Môn'] == 'Tổng cộng']
                        if not row_total.empty:
                            tiet_giangday_hk1_qdthieu = row_total['QĐ Thiếu'].values[0]
                            tiet_giangday_hk1_qdthua = row_total['QĐ thừa'].values[0]
                    if df_hk2 is not None and not df_hk2.empty:
                        row_total = df_hk2[df_hk2['Lớp // Môn'] == 'Tổng cộng']
                        if not row_total.empty:
                            tiet_giangday_hk2_qdthieu = row_total['QĐ Thiếu'].values[0]
                            tiet_giangday_hk2_qdthua = row_total['QĐ thừa'].values[0]
                    if tiet_giangday_hk1_qdthieu == 0 and len(dfs) > 0 and 'QĐ Thiếu' in dfs[0]:
                        tiet_giangday_hk1_qdthieu = dfs[0]['QĐ Thiếu'].sum()
                    if tiet_giangday_hk1_qdthua == 0 and len(dfs) > 0 and 'QĐ thừa' in dfs[0]:
                        tiet_giangday_hk1_qdthua = dfs[0]['QĐ thừa'].sum()
                    if tiet_giangday_hk2_qdthieu == 0 and len(dfs) > 0 and 'QĐ Thiếu' in dfs[0]:
                        tiet_giangday_hk2_qdthieu = dfs[0]['QĐ Thiếu'].sum()
                    if tiet_giangday_hk2_qdthua == 0 and len(dfs) > 0 and 'QĐ thừa' in dfs[0]:
                        tiet_giangday_hk2_qdthua = dfs[0]['QĐ thừa'].sum()
                    ra_de_cham_thi_hk1 = 0
                    ra_de_cham_thi_hk2 = 0
                    if len(dfs) > 1:
                        df_thi = dfs[1]
                        if 'Học kỳ 1 (Tiết)' in df_thi.columns:
                            ra_de_cham_thi_hk1 = pd.to_numeric(df_thi['Học kỳ 1 (Tiết)'], errors='coerce').sum()
                        elif 'Tiết quy đổi HK1' in df_thi.columns:
                            ra_de_cham_thi_hk1 = pd.to_numeric(df_thi['Tiết quy đổi HK1'], errors='coerce').sum()
                        if 'Học kỳ 2 (Tiết)' in df_thi.columns:
                            ra_de_cham_thi_hk2 = pd.to_numeric(df_thi['Học kỳ 2 (Tiết)'], errors='coerce').sum()
                        elif 'Tiết quy đổi HK2' in df_thi.columns:
                            ra_de_cham_thi_hk2 = pd.to_numeric(df_thi['Tiết quy đổi HK2'], errors='coerce').sum()
                    giam_gio = 0
                    if len(dfs) > 2:
                        df_giam = dfs[2]
                        if 'Tổng tiết' in df_giam.columns:
                            giam_gio = pd.to_numeric(df_giam['Tổng tiết'], errors='coerce').sum()
                        elif 'Số tiết giảm' in df_giam.columns:
                            giam_gio = pd.to_numeric(df_giam['Số tiết giảm'], errors='coerce').sum()
                    hoatdong_nckh = 0
                    hoatdong_thuctap = 0
                    hoatdong_khac = 0
                    if len(dfs) > 3 and not dfs[3].empty:
                        df_hd = dfs[3]
                        if 'MÃ NCKH' in df_hd.columns and 'Giờ quy đổi' in df_hd.columns:
                            hoatdong_nckh = df_hd.loc[df_hd['MÃ NCKH'] == 'NCKH', 'Giờ quy đổi'].sum()
                        if 'Mã HĐ' in df_hd.columns and 'Giờ quy đổi' in df_hd.columns:
                            hoatdong_thuctap = df_hd.loc[df_hd['Mã HĐ'] == 'HD07', 'Giờ quy đổi'].sum()
                        if 'MÃ NCKH' in df_hd.columns and 'Giờ quy đổi' in df_hd.columns:
                            hoatdong_khac = df_hd.loc[df_hd['MÃ NCKH'] == 'BT', 'Giờ quy đổi'].sum()
                    tong_thuchien_du = tiet_giangday_hk1_qdthua + tiet_giangday_hk2_qdthua + ra_de_cham_thi_hk1 + ra_de_cham_thi_hk2 + hoatdong_nckh + hoatdong_thuctap + hoatdong_khac - giam_gio
                    tong_thuchien_thieu = tiet_giangday_hk1_qdthieu + tiet_giangday_hk2_qdthieu + ra_de_cham_thi_hk1 + ra_de_cham_thi_hk2 + hoatdong_nckh + hoatdong_thuctap + hoatdong_khac - giam_gio
                    du_gio = max(0, tong_thuchien_du - giochuan)
                    thieu_gio = max(0, giochuan - tong_thuchien_thieu)
                    chuangv = st.session_state.get('chuan_gv', 'CĐ')
                    if chuangv in ['CĐ', 'TC']:
                        giochuan = 594
                    elif chuangv in ['CĐMC', 'TCMC']:
                        giochuan = 616
                    else:
                        giochuan = 594
                    if chuangv in ['CĐ', 'CĐMC']:
                        dinhmuc_giangday = giochuan / 44 * 32
                    elif chuangv in ['TC', 'TCMC']:
                        dinhmuc_giangday = giochuan / 44 * 36
                    else:
                        dinhmuc_giangday = giochuan / 44 * 32
                    if chuangv in ['CĐ', 'CĐMC']:
                        dinhmuc_nckh = giochuan / 44 * 8
                    elif chuangv in ['TC', 'TCMC']:
                        dinhmuc_nckh = giochuan / 44 * 4
                    else:
                        dinhmuc_nckh = giochuan / 44 * 8
                    dinhmuc_thuctap = giochuan / 44 * 4
                    dinhmuc_list = [round(dinhmuc_giangday, 2), '', '', '', '', '', round(dinhmuc_nckh, 2), round(dinhmuc_thuctap, 2), '']
                    dinhmuc_tongcong = sum([v for v in dinhmuc_list if isinstance(v, (int, float)) and v != ''])
                    dinhmuc_list.append(round(dinhmuc_tongcong, 2))
                    quydoi_du_list = ["", tiet_giangday_hk1_qdthua, tiet_giangday_hk2_qdthua, ra_de_cham_thi_hk1, ra_de_cham_thi_hk2, giam_gio, hoatdong_nckh, hoatdong_thuctap, hoatdong_khac]
                    quydoi_thieu_list = ["", tiet_giangday_hk1_qdthieu, tiet_giangday_hk2_qdthieu, ra_de_cham_thi_hk1, ra_de_cham_thi_hk2, giam_gio, hoatdong_nckh, hoatdong_thuctap, hoatdong_khac]
                    quydoi_du_tongcong = sum([v for v in quydoi_du_list[1:] if isinstance(v, (int, float)) and v != ''])
                    quydoi_thieu_tongcong = sum([v for v in quydoi_thieu_list[1:] if isinstance(v, (int, float)) and v != ''])
                    quydoi_du_list.append(round(quydoi_du_tongcong, 2))
                    quydoi_thieu_list.append(round(quydoi_thieu_tongcong, 2))
                    muc_list = ["(1)", "(2)", "(3)", "(4)", "(5)", "(6)", "(7)", "(8)", "(9)", ""]
                    noidung_list = [
                        "Định mức giảng dạy của GV",
                        "Tiết giảng dạy quy đổi (HK1)",
                        "Tiết giảng dạy quy đổi (HK2)",
                        "Ra đề, Coi thi, Chấm thi (HK1)",
                        "Ra đề, Coi thi, Chấm thi (HK2)",
                        "Giảm giờ Kiêm nhiệm QLý,GVCN...",
                        "Học tập, bồi dưỡng,NCKH",
                        "Thực tập tại doanh nghiệp",
                        "HD chuyên môn khác quy đổi",
                        "Tổng cộng"
                    ]
                    data = {
                        "MỤC": muc_list,
                        "NỘI DUNG QUY ĐỔI": noidung_list,
                        "Định Mức": dinhmuc_list,
                        "Quy đổi (Dư giờ)": quydoi_du_list,
                        "Quy đổi (Thiếu giờ)": quydoi_thieu_list
                    }
                    df_tonghop = pd.DataFrame(data)
                    df_tonghop = df_tonghop.where(pd.notnull(df_tonghop), '')
                    def zero_to_blank(val):
                        if val == 0 or val == 0.0:
                            return ''
                        return val
                    df_tonghop = df_tonghop.applymap(zero_to_blank)
                    return df_tonghop
                df_tonghop = build_bang_tonghop(dfs, giochuan)
                st.dataframe(df_tonghop.style.format(precision=1).set_properties(**{'text-align': 'center'}), use_container_width=True)
                st.session_state['df_all_tonghop'] = df_tonghop
                st.session_state['export_ready'] = True
            if not found_any:
                st.warning("Không có dữ liệu nào để tổng hợp từ các sheet 'output_'.")
        except Exception as e:
            st.error(f"Lỗi khi tải dữ liệu từ Google Sheet: {e}")
    # Chỉ hiển thị nút tải Excel mẫu khi đã tổng hợp xong
    export_ready = st.session_state.get('export_ready', False)
    if export_ready:
        import os
        template_path = os.path.join('data_base', 'mau_kegio.xlsx')
        spreadsheet = st.session_state.get('spreadsheet')
        df_mon = st.session_state.get('df_mon', pd.DataFrame())
        success, result = export_giangday_to_excel(spreadsheet=spreadsheet, df_mon=df_mon, template_path=template_path)
        if success:
            with open(result, 'rb') as f:
                st.download_button('Tải file Excel kết quả', f, file_name='ke_khai_gio_day.xlsx')
            st.success('Đã xuất dữ liệu ra file Excel mẫu!')
        else:
            st.error(result)

def main():
    tonghop_ketqua()

if __name__ == "__main__":
    main()
