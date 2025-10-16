import streamlit as st
import pandas as pd
import numpy as np
import io
from typing import List, Tuple, Dict, Any

st.set_page_config(layout="wide", page_title="Kê khai giờ dạy - Admin")
st.title("Kê khai giờ dạy - Admin")

st.markdown("""
### 1. Upload file Excel dữ liệu môn học
""")
# --- Load các bảng dữ liệu nền từ session_state ---
df_giaovien = st.session_state.get('df_giaovien', pd.DataFrame())
df_khoa = st.session_state.get('df_khoa', pd.DataFrame())
df_lop_g = st.session_state.get('df_lop', pd.DataFrame())
df_mon = st.session_state.get('df_mon', pd.DataFrame())
df_ngaytuan_g = st.session_state.get('df_ngaytuan', pd.DataFrame())
df_hesosiso_g = st.session_state.get('df_hesosiso', pd.DataFrame())
df_lopghep_g = st.session_state.get('df_lopghep', pd.DataFrame())
df_loptach_g = st.session_state.get('df_loptach', pd.DataFrame())
df_lopsc_g = st.session_state.get('df_lopsc', pd.DataFrame())
chuangv = st.session_state.get('chuangv', "Cao đẳng")  # Hoặc lấy từ input

def tao_cac_bang_he_so() -> dict:
    """Tạo và trả về một từ điển chứa tất cả các bảng hệ số."""
    data_cd = {
        'Môn_MC': [1.00, 0.89, 0.79, 1.00],
        'Môn_MĐ/MH': [1.00, 0.89, 0.79, 1.00],
        'Môn_VH': [1.00, 1.00, 1.00, 1.00]
    }
    df_cd = pd.DataFrame(data_cd, index=['Lớp_CĐ', 'Lớp_TC', 'Lớp_SC', 'Lớp_VH'])

    data_cdmc = {
        'Môn_MC': [1.00, 0.88, 0.79, 1.00],
        'Môn_MĐ/MH': [1.00, 0.89, 0.79, 1.00],
        'Môn_VH': [1.00, 1.00, 1.00, 1.00]
    }
    df_cdmc = pd.DataFrame(data_cdmc, index=['Lớp_CĐ', 'Lớp_TC', 'Lớp_SC', 'Lớp_VH'])

    data_tc = {
        'Môn_MC': [1.00, 1.00, 0.89, 1.00],
        'Môn_MĐ/MH': [1.00, 1.00, 0.89, 1.00],
        'Môn_VH': [1.00, 1.00, 1.00, 1.00]
    }
    df_tc = pd.DataFrame(data_tc, index=['Lớp_CĐ', 'Lớp_TC', 'Lớp_SC', 'Lớp_VH'])

    data_tcmc = {
        'Môn_MC': [1.00, 1.00, 0.89, 1.00],
        'Môn_MĐ/MH': [1.00, 1.00, 0.89, 1.00],
        'Môn_VH': [1.00, 1.00, 1.00, 1.00]
    }
    df_tcmc = pd.DataFrame(data_tcmc, index=['Lớp_CĐ', 'Lớp_TC', 'Lớp_SC', 'Lớp_VH'])

    data_vh = {
        'Môn_MC': [1.00, 1.00, 1.00, 1.00],
        'Môn_MĐ/MH': [1.00, 1.00, 1.00, 1.00],
        'Môn_VH': [1.00, 1.00, 1.00, 1.00]
    }
    df_vh = pd.DataFrame(data_vh, index=['Lớp_CĐ', 'Lớp_TC', 'Lớp_SC', 'Lớp_VH'])

    return {
        'CĐ': df_cd,
        'CĐMC': df_cdmc,
        'TC': df_tc,
        'TCMC': df_tcmc,
        'VH': df_vh
    }
def tra_cuu_heso_tccd(mamon_nganh: str, chuan_gv: str) -> float:
    """
    Tra cứu hệ số TC/CĐ dựa vào mã môn ngành và chuẩn GV.
    """
    #st.write(mamon_nganh)

    bang_he_so = tao_cac_bang_he_so()
    if chuan_gv not in bang_he_so:
        return 1.0  # Giá trị mặc định nếu không tìm thấy chuẩn GV
    loai_lop, loai_mon = phan_loai_ma_mon(mamon_nganh)
    try:
        return float(bang_he_so[chuan_gv].loc[loai_lop, loai_mon])
    except Exception:
        return 1.0
# Bước 2: Các hàm logic
def phan_loai_ma_mon(ma_mon: str) -> Tuple[str, str]:
    """Xác định loại lớp và loại môn cho một mã môn duy nhất."""
    ma_mon_upper = str(ma_mon).upper()
    #st.write(ma_mon_upper)
    # Xác định loại lớp
    ky_tu_dau = ma_mon_upper[0]
    if ky_tu_dau == '1':
        loai_lop = 'Lớp_CĐ'
    elif ky_tu_dau == '2':
        loai_lop = 'Lớp_TC'
    elif ky_tu_dau == '3':
        loai_lop = 'Lớp_SC'
    # Xác định loại môn
    if 'MC' in ma_mon_upper:
        loai_mon = 'Môn_MC'
    elif 'MH' in ma_mon_upper or 'MĐ' in ma_mon_upper:
        loai_mon = 'Môn_MĐ/MH'
    elif 'VH' in ma_mon_upper:
        loai_mon = 'Môn_VH'
    else:
        loai_mon = 'Không tìm thấy'
    #st.write(loai_mon,loai_lop)
    return loai_lop, loai_mon
def timhesomon_siso(siso, is_heavy_duty, lesson_type, df_hesosiso_g):
    """
    Tìm hệ số quy đổi dựa trên sĩ số, loại tiết (LT/TH) và điều kiện nặng nhọc.
    
    Tham số:
    - siso: Sĩ số của lớp học.
    - is_heavy_duty: True nếu môn học là nặng nhọc, False nếu bình thường.
    - lesson_type: 'LT' cho tiết Lý thuyết, 'TH' cho tiết Thực hành.
    - df_hesosiso_g: DataFrame chứa bảng tra cứu hệ số.
    """
    try:
        cleaned_siso = int(float(siso))
    except (ValueError, TypeError):
        cleaned_siso = 0
    siso = cleaned_siso

    df_hesosiso = df_hesosiso_g.copy()
    for col in ['LT min', 'LT max', 'TH min', 'TH max', 'THNN min', 'THNN max', 'Hệ số']:
        if col not in df_hesosiso.columns:
            return 1.0

    heso_siso = 1.0

    if lesson_type == 'LT':
        for _, row in df_hesosiso.iterrows():
            if row['LT min'] <= siso <= row['LT max']:
                heso_siso = row['Hệ số']
                break
    elif lesson_type == 'TH':
        if is_heavy_duty:
            for _, row in df_hesosiso.iterrows():
                if row['THNN min'] <= siso <= row['THNN max']:
                    heso_siso = row['Hệ số']
                    break
        else:
            for _, row in df_hesosiso.iterrows():
                if row['TH min'] <= siso <= row['TH max']:
                    heso_siso = row['Hệ số']
                    break
    return heso_siso
def get_arr_tiet_from_state(mon_state):
    cach_ke = mon_state.get('cach_ke', '')
    if cach_ke == 'Kê theo MĐ, MH':
        arr_tiet = [int(x) for x in str(mon_state.get('tiet', '')).split() if x]
        arr_tiet_lt = arr_tiet
        arr_tiet_th = [0]*len(arr_tiet)
    else:
        arr_tiet = [int(x) for x in str(mon_state.get('tiet', '')).split() if x]
        arr_tiet_lt = [int(x) for x in str(mon_state.get('tiet_lt', '0')).split() if x]
        arr_tiet_th = [int(x) for x in str(mon_state.get('tiet_th', '0')).split() if x]
    return arr_tiet, arr_tiet_lt, arr_tiet_th
def process_mon_data(row_input_data, df_lop_g, df_mon, df_ngaytuan_g, df_hesosiso_g):
    # Đảm bảo cột 'HS TC/CĐ' luôn tồn tại trước khi tính toán
    # Sử dụng trực tiếp loc_data_lop và loc_data_mon đã truyền vào, không cần lọc lại
    malop_info = df_lop_g  # df_lop_g lúc này là loc_data_lop
    mamon_info = df_mon    # df_mon lúc này là loc_data_mon
    # Kiểm tra số dòng, đảm bảo chỉ lấy đúng 1 dòng
    #st.write(mamon_info)
    if malop_info.empty:
        return pd.DataFrame(), {"error": "Không tìm thấy thông tin cho lớp."}
    if mamon_info.empty:
        return pd.DataFrame(), {"error": "Không tìm thấy thông tin cho môn học."}
    if len(malop_info) > 1:
        return pd.DataFrame(), {"error": "loc_data_lop có nhiều hơn 1 dòng, cần lọc chính xác."}
    if len(mamon_info) > 1:
        return pd.DataFrame(), {"error": "loc_data_mon có nhiều hơn 1 dòng, cần lọc chính xác."}
    # Lấy tiết từ các cột T1-T23 (HK1) hoặc T22-T48 (HK2) theo session_state['hocky']
    dulieu_info = row_input_data
    arr_tiet_list = []
    hocky = st.session_state.get('hocky', 'HK1')
    if hocky == 'HK1':
        tiet_start, tiet_end = 1, 23
    else:
        tiet_start, tiet_end = 22, 48
    for i in range(tiet_start, tiet_end + 1):
        tiet = dulieu_info.get(f'T{i}', 0)
        if tiet is None or (isinstance(tiet, float) and np.isnan(tiet)):
            tiet = 0
        try:
            arr_tiet_list.append(int(tiet))
        except:
            arr_tiet_list.append(0)
    tuanbatdau = None
    tuanketthuc = None
    # Xác định tuần thực tế dựa trên học kỳ
    if hocky == 'HK1':
        week_offset = 1
    else:
        week_offset = 22
    for idx, val in enumerate(arr_tiet_list):
        if val > 0:
            tuanbatdau = week_offset + idx
            break
    for idx in range(len(arr_tiet_list)-1, -1, -1):
        if arr_tiet_list[idx] > 0:
            tuanketthuc = week_offset + idx
            break
    if tuanbatdau is None or tuanketthuc is None:
        return pd.DataFrame(), {"error": "Không có tuần nào có dữ liệu tiết > 0."}
    
    st.write(f"Tuan bat dau: {tuanbatdau}, Tuan ket thuc: {tuanketthuc}")
    # Chỉ lấy dữ liệu tuần trong khoảng này (chuyển về chỉ số mảng)
    start_idx = tuanbatdau - week_offset
    end_idx = tuanketthuc - week_offset
    arr_tiet = arr_tiet_list[start_idx:end_idx+1]
    
    locdulieu_info = df_ngaytuan_g[(df_ngaytuan_g['Tuần'] >= tuanbatdau) & (df_ngaytuan_g['Tuần'] <= tuanketthuc)].copy()
    if len(locdulieu_info) != len(arr_tiet):
        return pd.DataFrame(), {"error": f"Số tuần ({len(locdulieu_info)}) không khớp số tiết ({len(arr_tiet)})."}
    #st.write(locdulieu_info)
        
    kieu_tinh_mdmh = mamon_info['Tính MĐ/MH'].iloc[0]
    loai_mon_mdmh = mamon_info['MH/MĐ'].iloc[0]
    # Xác định loại tiết
    if kieu_tinh_mdmh == 'LT':
        arr_tiet_lt = arr_tiet
        arr_tiet_th = np.zeros_like(arr_tiet)
    elif kieu_tinh_mdmh == 'TH':
        arr_tiet_lt = np.zeros_like(arr_tiet)
        arr_tiet_th = arr_tiet
    else:
        if loai_mon_mdmh == 'MH':
            arr_tiet_lt = arr_tiet
            arr_tiet_th = np.zeros_like(arr_tiet)
        elif loai_mon_mdmh == 'MĐ':
            arr_tiet_lt = np.zeros_like(arr_tiet)
            arr_tiet_th = arr_tiet
    # Xử lý dữ liệu đầu ra
    if 'Tháng' not in locdulieu_info.columns:
        found = False
        for col in locdulieu_info.columns:
            if col.lower().startswith('thang'):
                locdulieu_info = locdulieu_info.rename(columns={col: 'Tháng'})
                found = True
                break
        if not found:
            return pd.DataFrame(), {"error": "Không tìm thấy cột 'Tháng' trong dữ liệu tuần/ngày."}
    df_result = locdulieu_info[['Tuần', 'Từ ngày đến ngày']].copy()
    df_result.rename(columns={'Từ ngày đến ngày': 'Ngày'}, inplace=True)
    week_to_month = dict(zip(df_ngaytuan_g['Tuần'], df_ngaytuan_g['Tháng']))
    df_result['Tháng'] = df_result['Tuần'].map(week_to_month)
    siso_list = []
    for month in df_result['Tháng']:
        month_col = f"Tháng {month}"
        siso = malop_info[month_col].iloc[0] if month_col in malop_info.columns else 0
        siso_list.append(siso)
    df_result['Sĩ số'] = siso_list
    df_result['Tiết'] = arr_tiet
    df_result['Tiết_LT'] = arr_tiet_lt
    df_result['Tiết_TH'] = arr_tiet_th
    # Tính hệ số TC/CĐ cho từng dòng dựa vào mã ngành và chuẩn GV
    # mamon_nganh = "201Y_MH07"
    chuan_gv = chuan_gv_selected

    if 'Mã_môn_ngành' in mamon_info.columns:
        mamon_nganh = mamon_info['Mã_môn_ngành'].iloc[0]
    heso_tccd = tra_cuu_heso_tccd(mamon_nganh, chuan_gv)
    df_result['HS TC/CĐ'] = round(float(heso_tccd), 2)
    # Tra cứu hệ số sĩ số LT/TH từ bảng hệ số, giống logic quydoi_gioday.py
    # Xác định is_heavy_duty từ dữ liệu môn học
    is_heavy_duty = False
    if 'Nặng_nhọc' in mamon_info.columns:
        is_heavy_duty = mamon_info['Nặng_nhọc'].iloc[0] == 'NN'
    heso_lt_list, heso_th_list = [], []
    for siso in df_result['Sĩ số']:
        lt = timhesomon_siso(siso, is_heavy_duty, 'LT', df_hesosiso_g)
        th = timhesomon_siso(siso, is_heavy_duty, 'TH', df_hesosiso_g)
        heso_lt_list.append(lt)
        heso_th_list.append(th)
 
    df_result['HS_SS_LT'] = heso_lt_list
    df_result['HS_SS_TH'] = heso_th_list
    df_result["QĐ thừa"] = (df_result["Tiết_LT"] * df_result["HS_SS_LT"]) + (df_result["Tiết_TH"] * df_result["HS_SS_TH"])
    df_result["HS_SS_LT_tron"] = df_result["HS_SS_LT"].clip(lower=1)
    df_result["HS_SS_TH_tron"] = df_result["HS_SS_TH"].clip(lower=1)
    df_result["QĐ thiếu"] = df_result["HS TC/CĐ"] * ((df_result["Tiết_LT"] * df_result["HS_SS_LT_tron"]) + (df_result["HS_SS_TH_tron"] * df_result["Tiết_TH"]))
    rounding_map = {"Sĩ số": 0, "Tiết": 1, "HS_SS_LT": 1, "HS_SS_TH": 1, "QĐ thừa": 1, "QĐ thiếu": 1, "HS TC/CĐ": 2, "Tiết_LT": 1, "Tiết_TH": 1}
    #st.write("Bảng trước khi xóa cột")
    #st.write(df_result)
    for col, decimals in rounding_map.items():
        if col in df_result.columns:
            df_result[col] = pd.to_numeric(df_result[col], errors='coerce').fillna(0).round(decimals)
    final_columns = ["Tuần", "Ngày", "Tiết", "Sĩ số", "HS TC/CĐ", "Tiết_LT", "Tiết_TH", "HS_SS_LT", "HS_SS_TH", "QĐ thừa", "QĐ thiếu"]
    df_final = df_result[[col for col in final_columns if col in df_result.columns]]
    return df_final, {}

with st.expander("Lưu trữ dữ kiệu giảng dạy của giảng viên từ kết hoạch giảng dạy"):
    uploaded_file = st.file_uploader("Chọn file Excel nhập dữ liệu môn học", type=["xlsx", "xls"])
    if uploaded_file:
        # Đọc tên các sheet
        import openpyxl
        from openpyxl import load_workbook
        import os
        import tempfile
        # Lưu file tạm để openpyxl đọc sheet
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
            tmpfile.write(uploaded_file.read())
            tmpfile_path = tmpfile.name
        wb = load_workbook(tmpfile_path, read_only=True)
        sheet_names = wb.sheetnames
        # Chỉ lấy các sheet HK1/HK2 nếu có
        available_sheets = [s for s in sheet_names if s in ['HK1', 'HK2']]

        if not available_sheets:
            st.error("File Excel không có sheet HK1 hoặc HK2!")
            os.unlink(tmpfile_path)
        else:
            selected_sheet = st.selectbox("Chọn học kỳ để xử lý dữ liệu:", options=available_sheets)
            df_input = pd.read_excel(tmpfile_path, sheet_name=selected_sheet)
            os.unlink(tmpfile_path)
        st.session_state['hocky'] = selected_sheet
            # Nếu có cột ten_gv hoặc Ten_GV thì tạo danh sách giáo viên
        gv_col = None
        for col in ['ten_gv', 'Ten_GV']:
            if col in df_input.columns:
                gv_col = col
                break
        if gv_col:
            # Tạo list giáo viên dạng 'magv - ten_gv'
            gv_list_raw = df_input[gv_col].dropna().unique()
            gv_display_list = []
            gv_map = {}
            for ten_gv in gv_list_raw:
                magv = ''
                if not df_giaovien.empty and 'Tên giảng viên' in df_giaovien.columns:
                    matched_row = df_giaovien[df_giaovien['Tên giảng viên'] == ten_gv]
                    if not matched_row.empty and 'Magv' in matched_row.columns:
                        magv = str(matched_row['Magv'].iloc[0])
                display_name = f"{magv} - {ten_gv}" if magv else ten_gv
                gv_display_list.append(display_name)
                gv_map[display_name] = ten_gv
            selected_gv_display = st.selectbox("Chọn giáo viên để lọc dữ liệu", options=sorted(gv_display_list))
            selected_gv = gv_map[selected_gv_display]
            df_input = df_input[df_input[gv_col] == selected_gv].copy()
            # Ánh xạ selected_gv với df_giaovien để lấy Magv và Chức vụ_HT
            ma_gv = ''
            chuc_vu_ht = ''
            if not df_giaovien.empty and 'Tên giảng viên' in df_giaovien.columns:
                matched_row = df_giaovien[df_giaovien['Tên giảng viên'] == selected_gv]
                if not matched_row.empty:
                    if 'Magv' in matched_row.columns:
                        ma_gv = matched_row['Magv'].iloc[0]
                    if 'Chức vụ_HT' in matched_row.columns:
                        chuc_vu_ht = matched_row['Chức vụ_HT'].iloc[0]
            #st.write(f"Giáo viên đã chọn: {selected_gv}")
            st.write(f"Mã GV: {ma_gv}")
            #st.write(f"Chức vụ: {chuc_vu_ht}")
            # Lấy ký tự đầu tiên của Magv, ánh xạ với Mã_khoa của df_khoa (đảm bảo kiểu dữ liệu đúng)
            khoa_phong_trungtam = ''
            debug_info = {}
            if ma_gv and not df_khoa.empty and 'Mã_khoa' in df_khoa.columns and 'Khoa/Phòng/Trung tâm' in df_khoa.columns:
                try:
                    first_digit = int(str(ma_gv)[0])
                    # Đảm bảo cột Mã_khoa là int
                    df_khoa['Mã_khoa'] = pd.to_numeric(df_khoa['Mã_khoa'], errors='coerce').fillna(-1).astype(int)
                    debug_info['first_digit'] = first_digit
                    debug_info['df_khoa_Mã_khoa'] = df_khoa['Mã_khoa'].tolist()
                    matched_khoa = df_khoa[df_khoa['Mã_khoa'] == first_digit]
                    debug_info['matched_rows'] = matched_khoa.shape[0]
                    if not matched_khoa.empty:
                        khoa_phong_trungtam = matched_khoa['Khoa/Phòng/Trung tâm'].iloc[0]
                except Exception as e:
                    debug_info['error'] = str(e)
            #st.write(f"Khoa/Phòng/Trung tâm: {khoa_phong_trungtam}")
        # Chọn chuẩn giáo viên
        chuan_gv_selected = st.selectbox("Chuẩn GV", options=["CĐ", "CĐMC", "TC", "TCMC", "VH"], index=2)
        st.session_state['chuan_gv'] = chuan_gv_selected
        df_gv_info = pd.DataFrame([{
        "Mã_gv": ma_gv,
        "Tên_gv": selected_gv,
        "chucvu_hientai": chuc_vu_ht,
        "khoa": khoa_phong_trungtam,
        "chuan_gv": st.session_state['chuan_gv'],
        "gio_chuan": 0.0,
        "thongtin_giamgio": "",
        }])
        st.session_state['df_gv_info'] = df_gv_info
        # Chuẩn hóa tên cột về đúng định dạng code sử dụng
        col_map = {
            'Ten_gv': 'Ten_GV',
            'ten_gv': 'Ten_GV',
            'Mon_hoc': 'mon_hoc',
            'mon_hoc': 'mon_hoc',
            'Lop_hoc': 'lop_hoc',
            'lop_hoc': 'lop_hoc'
        }
        df_input.rename(columns={k: v for k, v in col_map.items() if k in df_input.columns}, inplace=True)
        # Thay thế toàn bộ None/NaN thành 0 cho các cột tiết theo học kỳ 1 - 23 và 22 - 48
        tiet_start, tiet_end = (1, 23) if st.session_state.get('hocky') == 'HK1' else (22, 48)
        for i in range(tiet_start, tiet_end + 1):
            col = f'T{i}'
            if col in df_input.columns:
                df_input[col] = df_input[col].fillna(0)
        # Đảm bảo cột Ten_GV nằm trước lop_hoc
        if 'Ten_GV' in df_input.columns:
            cols = list(df_input.columns)
            if cols[0] != 'Ten_GV':
                cols.remove('Ten_GV')
                cols.insert(0, 'Ten_GV')
                df_input = df_input[cols]
        #st.success("Đã upload dữ liệu. Xem trước bảng dữ liệu:")
        #st.dataframe(df_lop_g)
        # Hiển thị danh sách môn học phù hợp với từng lớp đã nhập
        #st.info("Danh sách môn học phù hợp với từng lớp đã nhập:")
        if 'lop_hoc' in df_input.columns:
            from difflib import get_close_matches
            # Thay thế các ký hiệu N1/N2/N3/Ca1/Ca2/Ca3 thành _1/_2/_3 trước khi xóa khoảng trắng
            def fix_lop_name(name):
                name = str(name)
                # Nếu có 'CĐ' thì tìm tên lớp gần đúng, không xóa khoảng trắng
                if 'CĐ' in name:
                    lop_hop_le_list = df_lop_g['Lớp'].astype(str).tolist() if not df_lop_g.empty and 'Lớp' in df_lop_g.columns else []
                    from difflib import get_close_matches
                    match = get_close_matches(name, lop_hop_le_list, n=1, cutoff=0.7)
                    return match[0] if match else name
                # Nếu không có 'CĐ', thực hiện chuẩn hóa như cũ
                for i in range(1, 5):
                    name = name.replace(f"(N{i})", f"_{i}")
                    name = name.replace(f"(n{i})", f"_{i}")
                    name = name.replace(f"(Ca{i})", f"_{i}")
                    name = name.replace(f"(ca{i})", f"_{i}")
                for pat, rep in [(" N1", "_1"), (" n1", "_1"), (" Ca1", "_1"), (" ca1", "_1"),
                                (" N2", "_2"), (" n2", "_2"), (" Ca2", "_2"), (" ca2", "_2"),
                                (" N3", "_3"), (" n3", "_3"), (" Ca3", "_3"), (" ca3", "_3"),
                                (" N4", "_4"), (" n4", "_4"), (" Ca4", "_4"), (" ca4", "_4")]:
                    name = name.replace(pat, rep)
                return name

            df_input['lop_hoc'] = df_input['lop_hoc'].apply(fix_lop_name)
            # Nếu không có 'CĐ' trong tên lớp thì mới loại bỏ khoảng trắng
            def remove_space_if_no_cd(name):
                name_str = str(name)
                if 'CĐ' in name_str:
                    return name_str.strip()
                else:
                    return name_str.replace(' ', '').strip()
            df_input['lop_hoc'] = df_input['lop_hoc'].apply(remove_space_if_no_cd)
            
            lop_hop_le_set = set(df_lop_g['Lớp']) if not df_lop_g.empty and 'Lớp' in df_lop_g.columns else set()
            for lop in df_input['lop_hoc'].drop_duplicates():
                #st.write(f"Lớp: {lop}")
                malop_info = df_lop_g[df_lop_g['Lớp'] == lop]
                if lop not in lop_hop_le_set:
                    st.error(f"Lớp '{lop}' không có trong danh sách lớp hợp lệ!")
                    continue
                if not malop_info.empty:
                    # ...existing code...
                    ma_lop = str(malop_info['Mã_lớp'].iloc[0]) if 'Mã_lớp' in malop_info.columns else ''
                    dsmon_code = ''
                    if len(ma_lop) >= 6:
                        A = ma_lop[2:5]
                        B = ma_lop[0:2]
                        last_char = ma_lop[-1]
                        if last_char == 'X':
                            dsmon_code = f"{A}X"
                        else:
                            try:
                                B_num = int(B)
                            except:
                                B_num = 0
                            if B_num >= 49:
                                dsmon_code = f"{A}Z"
                            else:
                                dsmon_code = f"{A}Y"
                    if dsmon_code and 'Mã_ngành' in df_mon.columns and 'Môn_học' in df_mon.columns:
                        mon_list = df_mon[df_mon['Mã_ngành'] == dsmon_code]['Môn_học'].dropna().astype(str).tolist()
                        if 'mon_list_by_lop' not in st.session_state:
                            st.session_state['mon_list_by_lop'] = {}
                        st.session_state['mon_list_by_lop'][lop] = mon_list
                        data_mon_list = df_mon[df_mon['Mã_ngành'] == dsmon_code].copy()
                        if 'data_mon_list_by_lop' not in st.session_state:
                            st.session_state['data_mon_list_by_lop'] = {}
                        st.session_state['data_mon_list_by_lop'][lop] = data_mon_list
                    else:
                        mon_list = []
                    mon_hoc_excel = df_input[df_input['lop_hoc'] == lop]['mon_hoc'].dropna().astype(str).tolist()
                    fuzzy_map = {}
                    need_fuzzy = False
                    for mh in mon_hoc_excel:
                        if mh not in mon_list:
                            match = get_close_matches(mh, mon_list, n=1, cutoff=0.6)
                            fuzzy_map[mh] = match[0] if match else ''
                            need_fuzzy = True
                    if need_fuzzy:
                        st.dataframe(pd.DataFrame({'Môn học nhập': list(fuzzy_map.keys()), 'Môn học gần đúng': list(fuzzy_map.values())}))

        output_rows = []
        lop_hop_le = set(df_lop_g['Lớp']) if not df_lop_g.empty and 'Lớp' in df_lop_g.columns else set()
        loi_lop = []
        if 'lop_hoc' not in df_input.columns:
            st.error("File Excel của bạn thiếu cột 'lop_hoc'. Vui lòng kiểm tra lại tên cột hoặc chuẩn hóa đúng định dạng!")
            st.info(f"Các cột hiện có trong file Excel:")
            st.dataframe(pd.DataFrame({'Tên cột': list(df_input.columns)}))
        else:
            from difflib import get_close_matches
            debug_rows = []
            # Tạo bản sao dữ liệu đầu vào để thay thế môn học gần đúng
            df_input_new = df_input.copy()
            for lop in df_input_new['lop_hoc'].drop_duplicates():
                malop_info = df_lop_g[df_lop_g['Lớp'] == lop]
                if not malop_info.empty:
                    # Lấy mon_list cho lớp này từ session_state
                    mon_list = st.session_state.get('mon_list_by_lop', {}).get(lop, [])
                    # Tìm giá trị gần đúng nhất trong mon_list so với từng giá trị mon_hoc đã nhập
                    mon_hoc_excel = df_input_new[df_input_new['lop_hoc'] == lop]['mon_hoc'].dropna().astype(str).tolist()
                    fuzzy_map = {}
                    for mh in mon_hoc_excel:
                        match = get_close_matches(mh, mon_list, n=1, cutoff=0.6)
                        fuzzy_map[mh] = match[0] if match else mh
                    # Thay thế vào df_input_new (đúng phạm vi, đúng mon_list)
                    mask = df_input_new['lop_hoc'] == lop
                    df_input_new.loc[mask, 'mon_hoc'] = df_input_new.loc[mask, 'mon_hoc'].apply(lambda x: fuzzy_map.get(str(x), x))
                    # Lưu fuzzy_map cho từng lớp để sử dụng sau này
                    if 'fuzzy_map_by_lop' not in st.session_state:
                        st.session_state['fuzzy_map_by_lop'] = {}
                    st.session_state['fuzzy_map_by_lop'][lop] = fuzzy_map
            # Tiếp tục kiểm tra và tính toán với dữ liệu đã thay thế
            #st.write(df_input_new)
            for idx, row in df_input_new.iterrows():
                ten_lop = row['lop_hoc'] if 'lop_hoc' in row else row.get('lop_hoc')
                # Lấy tên môn học gần đúng từ fuzzy_map_by_lop nếu có
                fuzzy_map = st.session_state.get('fuzzy_map_by_lop', {}).get(ten_lop, {})
                ten_mon_goc = row['mon_hoc'] if 'mon_hoc' in row else row.get('mon_hoc')
                ten_mon = fuzzy_map.get(str(ten_mon_goc), ten_mon_goc)
                # Lấy dữ liệu môn học đã lọc gần đúng cho lớp này
                loc_data_monhoc = None
                st.write(f"Xử lý dòng {idx}: Lớp '{ten_lop}', Môn '{ten_mon}' (gốc: '{ten_mon_goc}')")
                if 'data_mon_list_by_lop' in st.session_state and ten_lop in st.session_state['data_mon_list_by_lop']:
                    df_data_mon = st.session_state['data_mon_list_by_lop'][ten_lop]
                    loc_data_monhoc = df_data_mon[df_data_mon['Môn_học'] == ten_mon]
                # Lọc dữ liệu lớp theo tên lớp
                loc_data_lop = df_lop_g[df_lop_g['Lớp'] == ten_lop]
                
                # Nếu cần kiểm tra hoặc sử dụng loc_data_monhoc, có thể thêm xử lý tại đây
                debug_info = {'row': idx, 'lop_hoc': ten_lop, 'mon_hoc': ten_mon, 'status': '', 'detail': ''}
                # Lấy mon_list từ session_state nếu có
                mon_list = st.session_state.get('mon_list_by_lop', {}).get(ten_lop, [])
                debug_info['mon_hoc_hople'] = ', '.join(mon_list)
                if ten_lop not in lop_hop_le:
                    loi_lop.append(ten_lop)
                    debug_info['status'] = 'Lớp không hợp lệ'
                    debug_info['detail'] = f"Tên lớp '{ten_lop}' không có trong danh sách lớp hợp lệ."
                    debug_rows.append(debug_info)
                    continue
                #st.write(ten_mon, mon_list)
                ten_mon_norm = str(ten_mon).strip().lower()
                mon_list_norm = [str(m).strip().lower() for m in mon_list]
                found = any(ten_mon_norm == m for m in mon_list_norm)
                if not found:
                    debug_info['status'] = 'Môn học không hợp lệ'
                    debug_info['detail'] = f"Tên môn '{ten_mon}' không có trong danh sách môn học hợp lệ cho lớp '{ten_lop}'."
                    debug_rows.append(debug_info)
                    continue
                st.write(loc_data_monhoc)
                bangtonghop_mon, info = process_mon_data(row, loc_data_lop, loc_data_monhoc, df_ngaytuan_g, df_hesosiso_g)
                #st.write(bangtonghop_mon)
                if bangtonghop_mon.empty:
                    st.error(f"[ROW {idx}] Không tạo được bảng tổng hợp. Lý do: {info.get('error', 'Không rõ')}")
                else:
                    # Chèn cột Lớp_học và Mã_môn_ngành vào sau cột Tuần
                    bangtonghop_mon.insert(1, 'Lớp_học', ten_lop)
                    mamon_nganh = ''
                    if loc_data_monhoc is not None and not loc_data_monhoc.empty and 'Mã_môn_ngành' in loc_data_monhoc.columns:
                        mamon_nganh = loc_data_monhoc['Mã_môn_ngành'].iloc[0]
                    bangtonghop_mon.insert(2, 'Mã_môn_ngành', mamon_nganh)
                    output_rows.append(bangtonghop_mon)
            if output_rows:
                bangtonghop_all = pd.concat(output_rows, ignore_index=True)
                # Đổi tên cột 'Mã_môn_ngành' thành 'Mã_Môn_Ngành' nếu tồn tại
                if 'Mã_môn_ngành' in bangtonghop_all.columns:
                    bangtonghop_all.rename(columns={'Mã_môn_ngành': 'Mã_Môn_Ngành'}, inplace=True)
                # Tạo bản hiển thị thêm cột Tên_môn sau cột Lớp_học
                bangtonghop_display = bangtonghop_all.copy()
                if 'Lớp_học' in bangtonghop_display.columns:
                    idx = bangtonghop_display.columns.get_loc('Lớp_học')
                    if 'Tên_môn' in bangtonghop_display.columns:
                        pass
                    elif 'Mã_Môn_Ngành' in bangtonghop_display.columns:
                        # Tạo Series ánh xạ tên môn từ df_mon
                        ma_mon_map = df_mon.set_index('Mã_môn_ngành')['Môn_học'] if ('Mã_môn_ngành' in df_mon.columns and 'Môn_học' in df_mon.columns) else None
                        if ma_mon_map is not None:
                            ten_mon_series = bangtonghop_display['Mã_Môn_Ngành'].map(ma_mon_map).fillna('')
                            bangtonghop_display.insert(idx+1, 'Tên_môn', ten_mon_series)
                        else:
                            bangtonghop_display.insert(idx+1, 'Tên_môn', '')
                    else:
                        bangtonghop_display.insert(idx+1, 'Tên_môn', '')
            
            if loi_lop:
                st.error(f"Các tên lớp sau không hợp lệ: {', '.join([str(x) for x in loi_lop if pd.notna(x)])}")
                st.info(f"Các tên lớp bạn đã nhập trong file Excel:")
                st.dataframe(pd.DataFrame({'Tên lớp nhập': df_input['lop_hoc'].drop_duplicates().tolist()}))
                st.info(f"Vui lòng chọn đúng tên lớp trong danh sách sau:")
                st.dataframe(pd.DataFrame({'Lớp hợp lệ': sorted(lop_hop_le)}))
                if debug_rows:
                    st.info("Chi tiết kiểm tra từng dòng:")
                    st.dataframe(pd.DataFrame(debug_rows))
            elif output_rows:
                st.markdown("### 2. Kết quả xử lý - Xuất file Excel")
                # Ẩn cột Ngày và Mã_Môn_Ngành khi hiển thị
                cols_to_hide = ['Ngày', 'Mã_Môn_Ngành']
                cols_to_show = [col for col in bangtonghop_display.columns if col not in cols_to_hide]
                st.dataframe(bangtonghop_display[cols_to_show])

                # Nút lưu dữ liệu vào session_state và Google Sheet
                if st.button("Lưu dữ liệu"):
                    st.session_state['bangtonghop_all'] = bangtonghop_all.copy()
                    st.success("Đã lưu dữ liệu vào session_state['bangtonghop_all']!")
                    try:
                        import gspread
                        from google.oauth2.service_account import Credentials
                        from googleapiclient.discovery import build
                        creds_dict = st.secrets["gcp_service_account"]
                        folder_id = st.secrets["google_sheet"]["target_folder_id"]
                        template_file_id = st.secrets["google_sheet"]["template_file_id"]
                        ma_gv_sheet = str(ma_gv) if 'ma_gv' in locals() or 'ma_gv' in globals() else "output_giangday"
                        scopes = [
                            "https://www.googleapis.com/auth/drive",
                            "https://www.googleapis.com/auth/spreadsheets"
                        ]
                        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
                        gc = gspread.authorize(creds)
                        drive_service = build('drive', 'v3', credentials=creds)
                        query = f"name='{ma_gv_sheet}' and '{folder_id}' in parents and mimeType='application/vnd.google-apps.spreadsheet'"
                        results = drive_service.files().list(q=query, fields="files(id, name)").execute()
                        files = results.get('files', [])

                        # Xác định tên sheet cần lưu dựa vào học kỳ
                        hocky = st.session_state.get('hocky', 'HK1')
                        sheet_title = 'output_giangday' if hocky == 'HK1' else 'output_giangday(HK2)'
                        if files:
                            sheet_id = files[0]['id']
                            sh = gc.open_by_key(sheet_id)
                            worksheet = None
                            try:
                                worksheet = sh.worksheet(sheet_title)
                            except:
                                worksheet = sh.add_worksheet(title=sheet_title, rows=100, cols=20)
                            # Thêm hoặc cập nhật worksheet thongtin_gv
                            try:
                                ws_gv = sh.worksheet('thongtin_gv')
                            except:
                                ws_gv = sh.add_worksheet(title='thongtin_gv', rows=10, cols=20)
                            df_gv_info = st.session_state.get('df_gv_info')
                            if df_gv_info is not None:
                                ws_gv.clear()
                                ws_gv.update([df_gv_info.columns.values.tolist()] + df_gv_info.values.tolist())
                        else:
                            copied_file = drive_service.files().copy(
                                fileId=template_file_id,
                                body={"name": ma_gv_sheet, "parents": [folder_id]}
                            ).execute()
                            sheet_id = copied_file['id']
                            sh = gc.open_by_key(sheet_id)
                            worksheet = None
                            try:
                                worksheet = sh.worksheet(sheet_title)
                            except:
                                worksheet = sh.add_worksheet(title=sheet_title, rows=100, cols=20)
                            try:
                                ws_gv = sh.worksheet('thongtin_gv')
                            except:
                                ws_gv = sh.add_worksheet(title='thongtin_gv', rows=10, cols=20)
                            df_gv_info = st.session_state.get('df_gv_info')
                            if df_gv_info is not None:
                                ws_gv.clear()
                                ws_gv.update([df_gv_info.columns.values.tolist()] + df_gv_info.values.tolist())
                        worksheet.clear()
                        worksheet.update([bangtonghop_all.columns.values.tolist()] + bangtonghop_all.values.tolist())
                        st.success(f"Đã lưu dữ liệu vào Google Sheet: https://docs.google.com/spreadsheets/d/{sheet_id}")
                    except Exception as e:
                        st.error(f"Lỗi lưu Google Sheet: {e}")

with st.expander("Tạo file Excel tải về cho giảng viên"):

    import gspread
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds_dict = st.secrets["gcp_service_account"]
    folder_id = st.secrets["google_sheet"]["target_folder_id"]
    template_file_id = st.secrets["google_sheet"]["template_file_id"]
    scopes = [
        "https://www.googleapis.com/auth/drive",
        "https://www.googleapis.com/auth/spreadsheets"
    ]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    gc = gspread.authorize(creds)
    drive_service = build('drive', 'v3', credentials=creds)
    # Lấy danh sách file sheet trong folder
    query = f"'{folder_id}' in parents and mimeType='application/vnd.google-apps.spreadsheet'"
    results = drive_service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get('files', [])
    sheet_names = [f['name'] for f in files]
    # Ánh xạ sheet_names từ ma_gv vào cột Magv của df_giaovien để lấy giá trị cột Tên giảng viên
    magv_to_tengv = {}
    if not df_giaovien.empty and 'Magv' in df_giaovien.columns and 'Tên giảng viên' in df_giaovien.columns:
        for _, row in df_giaovien.iterrows():
            magv = str(row['Magv'])
            tengv = row['Tên giảng viên']
            magv_to_tengv[magv] = tengv
    # Tạo DataFrame ánh xạ sheet_names sang tên giảng viên
    df_sheet_map = pd.DataFrame({
        'Tên file Google Sheet': sheet_names,
        'Tên giảng viên': [magv_to_tengv.get(name, '') for name in sheet_names]
    })
    #st.write("Ánh xạ tên file Google Sheet sang tên giảng viên:")
    #st.dataframe(df_sheet_map)
    
    # Tạo list tên giảng viên từ ánh xạ
    # Tạo list giáo viên dạng 'magv - ten_gv' cho selectbox tải dữ liệu
    gv_display_list = []
    gv_map = {}
    for idx, row in df_sheet_map.iterrows():
        magv = ''
        ten_gv = row['Tên giảng viên']
        sheet_name = row['Tên file Google Sheet']
        # Loại bỏ entry nếu tên giảng viên trống
        if not ten_gv or str(ten_gv).strip() == '':
            continue
        if sheet_name in magv_to_tengv:
            magv = sheet_name
        display_name = f"{magv} - {ten_gv}" if magv else ten_gv
        gv_display_list.append(display_name)
        gv_map[display_name] = sheet_name
    selected_gv_display = st.selectbox("Chọn tên giảng viên để tải dữ liệu:", options=sorted(gv_display_list))
    selected_sheet_name = gv_map.get(selected_gv_display, None)
    if selected_sheet_name is None:
        st.error("Không tìm thấy file Google Sheet tương ứng với giáo viên đã chọn.")
        st.stop()
    # Khi chọn tên sheet, load dữ liệu từ file đó
    df_giangday = None
    df_giangday_hk2 = None
    df_gv_info = None
    def convert_comma_decimal(df, columns=None):
        import pandas as pd
        if df is None or df.empty:
            return df
        if columns is None:
            columns = df.select_dtypes(include='object').columns
        for col in columns:
            df[col] = df[col].apply(lambda x: float(str(x).replace(',', '.')) if isinstance(x, str) and ',' in x and str(x).replace(',', '.').replace('.', '', 1).isdigit() else x)
        return df
    if selected_sheet_name:
        file_obj = next((f for f in files if f['name'] == selected_sheet_name), None)
        if file_obj:
            sheet_id = file_obj['id']
            sh = gc.open_by_key(sheet_id)
            # Load worksheet output_giangday
            try:
                ws_giangday = sh.worksheet('output_giangday')
                records_giangday = ws_giangday.get_all_records()
                df_giangday = pd.DataFrame(records_giangday)
                df_giangday = convert_comma_decimal(df_giangday)
            except Exception as e:
                st.write(f"Không tìm thấy hoặc lỗi worksheet output_giangday: {e}")
                df_giangday = None
            # Load worksheet output_giangday(HK2)
            try:
                ws_giangday_hk2 = sh.worksheet('output_giangday(HK2)')
                records_giangday_hk2 = ws_giangday_hk2.get_all_records()
                df_giangday_hk2 = pd.DataFrame(records_giangday_hk2)
                df_giangday_hk2 = convert_comma_decimal(df_giangday_hk2)
            except Exception as e:
                st.write(f"Không tìm thấy hoặc lỗi worksheet output_giangday(HK2): {e}")
                df_giangday_hk2 = None
            # Load worksheet thongtin_gv
            try:
                ws_gv_info = sh.worksheet('thongtin_gv')
                records_gv_info = ws_gv_info.get_all_records()
                df_gv_info = pd.DataFrame(records_gv_info)
            except Exception as e:
                st.write(f"Không tìm thấy hoặc lỗi worksheet thongtin_gv: {e}")
                df_gv_info = None
    with st.expander("Xem thông tin giảng viên"):
        st.write(df_gv_info)
        st.session_state['df_gv_info'] = df_gv_info
    with st.expander("Xem dữ liệu giảng dạy HK1 từ Google Sheet"):
        # Tự động gom nhóm bảng HK1
        df_grouped_hk1 = None
        if df_giangday is not None and not df_giangday.empty:
            # Đảm bảo các cột cần thiết tồn tại
            required_cols = ['Lớp_học', 'Mã_Môn_Ngành', 'Tiết', 'QĐ thừa']
            missing_cols = [col for col in required_cols if col not in df_giangday.columns]
            if missing_cols:
                st.warning(f"Bảng giảng dạy HK1 thiếu các cột: {', '.join(missing_cols)}")
            else:
                # Gom nhóm theo Lớp_học và Mã_Môn_Ngành, tính tổng Tiết và QĐ thừa
                df_grouped_hk1 = df_giangday.groupby(['Lớp_học', 'Mã_Môn_Ngành'], as_index=False).agg({
                    'Tiết': 'sum',
                    'QĐ thừa': 'sum'
                })
                # Ánh xạ Môn học và MH/MĐ từ df_mon
                mon_map = df_mon.set_index('Mã_môn_ngành')['Môn_học'] if ('Mã_môn_ngành' in df_mon.columns and 'Môn_học' in df_mon.columns) else None
                mdmh_map = df_mon.set_index('Mã_môn_ngành')['MH/MĐ'] if ('Mã_môn_ngành' in df_mon.columns and 'MH/MĐ' in df_mon.columns) else None
                df_grouped_hk1['Môn_học'] = df_grouped_hk1['Mã_Môn_Ngành'].map(mon_map) if mon_map is not None else ''
                df_grouped_hk1['MH/MĐ'] = df_grouped_hk1['Mã_Môn_Ngành'].map(mdmh_map) if mdmh_map is not None else ''
                # Đổi tên cột QĐ thừa thành Tiết QĐ
                df_grouped_hk1.rename(columns={'QĐ thừa': 'Tiết QĐ'}, inplace=True)
                # Sắp xếp lại thứ tự cột
                cols_show = ['Lớp_học', 'Môn_học', 'Tiết', 'Tiết QĐ', 'MH/MĐ']
                df_grouped_hk1 = df_grouped_hk1[[col for col in cols_show if col in df_grouped_hk1.columns]]
                # Thêm dòng tổng cộng
                total_row = {
                    'Lớp_học': 'TỔNG',
                    'Môn_học': '',
                    'Tiết': df_grouped_hk1['Tiết'].sum(),
                    'Tiết QĐ': df_grouped_hk1['Tiết QĐ'].sum(),
                    'MH/MĐ': ''
                }
                df_grouped_hk1 = pd.concat([df_grouped_hk1, pd.DataFrame([total_row])], ignore_index=True)
                st.session_state['df_grouped_hk1'] = df_grouped_hk1.copy()
                st.write("#### Bảng tổng hợp giảng dạy HK1 (gộp theo lớp và môn)")
                st.dataframe(df_grouped_hk1)
        # Hiển thị bảng gốc như cũ
        #st.write("#### Bảng chi tiết HK1 từ Google Sheet")
        #st.write(df_giangday)
    with st.expander("Xem dữ liệu giảng dạy HK2 từ Google Sheet"):
        # Tự động gom nhóm bảng HK2
        df_grouped_hk2 = None
        if df_giangday_hk2 is not None and not df_giangday_hk2.empty:
            required_cols = ['Lớp_học', 'Mã_Môn_Ngành', 'Tiết', 'QĐ thừa']
            missing_cols = [col for col in required_cols if col not in df_giangday_hk2.columns]
            if missing_cols:
                st.warning(f"Bảng giảng dạy HK2 thiếu các cột: {', '.join(missing_cols)}")
            else:
                df_grouped_hk2 = df_giangday_hk2.groupby(['Lớp_học', 'Mã_Môn_Ngành'], as_index=False).agg({
                    'Tiết': 'sum',
                    'QĐ thừa': 'sum'
                })
                mon_map = df_mon.set_index('Mã_môn_ngành')['Môn_học'] if ('Mã_môn_ngành' in df_mon.columns and 'Môn_học' in df_mon.columns) else None
                mdmh_map = df_mon.set_index('Mã_môn_ngành')['MH/MĐ'] if ('Mã_môn_ngành' in df_mon.columns and 'MH/MĐ' in df_mon.columns) else None
                df_grouped_hk2['Môn_học'] = df_grouped_hk2['Mã_Môn_Ngành'].map(mon_map) if mon_map is not None else ''
                df_grouped_hk2['MH/MĐ'] = df_grouped_hk2['Mã_Môn_Ngành'].map(mdmh_map) if mdmh_map is not None else ''
                df_grouped_hk2.rename(columns={'QĐ thừa': 'Tiết QĐ'}, inplace=True)
                cols_show = ['Lớp_học', 'Môn_học', 'Tiết', 'Tiết QĐ', 'MH/MĐ']
                df_grouped_hk2 = df_grouped_hk2[[col for col in cols_show if col in df_grouped_hk2.columns]]
                # Thêm dòng tổng cộng
                total_row = {
                    'Lớp_học': 'TỔNG',
                    'Môn_học': '',
                    'Tiết': df_grouped_hk2['Tiết'].sum(),
                    'Tiết QĐ': df_grouped_hk2['Tiết QĐ'].sum(),
                    'MH/MĐ': ''
                }
                df_grouped_hk2 = pd.concat([df_grouped_hk2, pd.DataFrame([total_row])], ignore_index=True)
                st.session_state['df_grouped_hk2'] = df_grouped_hk2.copy()
                st.write("#### Bảng tổng hợp giảng dạy HK2 (gộp theo lớp và môn)")
                st.dataframe(df_grouped_hk2)
        #st.write("#### Bảng chi tiết HK2 từ Google Sheet")
        #st.write(df_giangday_hk2)
    # Tạo dummy spreadsheet để xuất file Excel
    class DummyWorksheet:
        def __init__(self, df, title):
            self._df = df
            self.title = title
        def get_all_records(self):
            import pandas as pd
            if self._df is None or not isinstance(self._df, pd.DataFrame) or self._df.empty:
                return []
            return self._df.to_dict(orient='records')
    class DummySpreadsheet:
        def __init__(self, df_giangday, df_giangday_hk2, df_gv_info):
            self._df_giangday = df_giangday
            self._df_giangday_hk2 = df_giangday_hk2
            self._df_gv_info = df_gv_info
        def worksheets(self):
            ws = [DummyWorksheet(self._df_giangday, 'output_giangday')]
            if self._df_giangday_hk2 is not None:
                ws.append(DummyWorksheet(self._df_giangday_hk2, 'output_giangday(HK2)'))
            if self._df_gv_info is not None:
                ws.append(DummyWorksheet(self._df_gv_info, 'thongtin_gv'))
            return ws
    spreadsheet = DummySpreadsheet(df_giangday, df_giangday_hk2, df_gv_info)
    from tonghop_kegio import export_giangday_to_excel
    #st.write("HK1:", spreadsheet._df_giangday)
    #st.write("HK2:", spreadsheet._df_giangday_hk2)
    #st.write("GV info:", spreadsheet._df_gv_info)
    template_path = 'data_base/mau_kegio.xlsx'
    ok, file_path = export_giangday_to_excel(spreadsheet=spreadsheet, df_mon=df_mon, template_path=template_path)
    #st.write("HK1:", df_giangday_hk1)
    #st.write("HK2:", df_giangday_hk2)
    if ok:
        ten_gv = ''
        ma_gv = ''
        if df_gv_info is not None:
            if 'Tên_gv' in df_gv_info.columns and not df_gv_info.empty:
                ten_gv = str(df_gv_info['Tên_gv'].iloc[0]).replace(' ', '_')
            if 'Mã_gv' in df_gv_info.columns and not df_gv_info.empty:
                ma_gv = str(df_gv_info['Mã_gv'].iloc[0])
        file_download_name = f"{ma_gv}_{ten_gv}.xlsx" if ma_gv and ten_gv else "output_giangday.xlsx"
        with open(file_path, 'rb') as f:
            st.download_button(
                label="Tải file Excel kết quả",
                data=f.read(),
                file_name=file_download_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        # Chỉ clear các biến session_state về giá trị mặc định để tránh xung đột khi chọn GV khác
        clear_defaults = {
            #'df_giangday': pd.DataFrame(),
            #'df_giangday_hk2': pd.DataFrame(),
            #'df_gv_info': pd.DataFrame(),
            'spreadsheet': None,
            #'df_mon': pd.DataFrame(),
            #'df_hk1': pd.DataFrame(),
            #'df_hk2': pd.DataFrame(),
            #'df_all_tonghop': pd.DataFrame(),
            #'export_ready': False,
            #'giochuan': 0,
            #'chuan_gv': ''
        }
        for key, val in clear_defaults.items():
            st.session_state[key] = val
    else:
        st.error(f"Lỗi xuất file Excel: {file_path}")
            # --- Upload file tonghop.xlsx và ghi tổng Tiết, Tiết QĐ vào cột 0, hàng theo tên giáo viên ---
    with st.expander("Cập nhật tổng Tiết và Tiết QĐ vào file tổng hợp"):
        uploaded_tonghop_file = st.file_uploader("Tải lên file tổng hợp (tonghop.xlsx)", type=["xlsx"])
        if uploaded_tonghop_file is not None:
            import openpyxl
            import tempfile
            # Đọc tổng Tiết và Tiết QĐ từ bảng HK1
            tong_tiet = None
            tong_tiet_qd = None
            df_grouped_hk1 = st.session_state.get('df_grouped_hk1')
            if df_grouped_hk1 is not None:
                tong_tiet = df_grouped_hk1['Tiết'].sum() if 'Tiết' in df_grouped_hk1.columns else None
                tong_tiet_qd = df_grouped_hk1['Tiết QĐ'].sum() if 'Tiết QĐ' in df_grouped_hk1.columns else None
            # Đọc tổng Tiết và Tiết QĐ từ bảng HK2
            tong_tiet_hk2 = None
            tong_tiet_qd_hk2 = None
            df_grouped_hk2 = st.session_state.get('df_grouped_hk2')
            if df_grouped_hk2 is not None:
                tong_tiet_hk2 = df_grouped_hk2['Tiết'].sum() if 'Tiết' in df_grouped_hk2.columns else None
                tong_tiet_qd_hk2 = df_grouped_hk2['Tiết QĐ'].sum() if 'Tiết QĐ' in df_grouped_hk2.columns else None
            # Lấy tên giáo viên
            ten_gv = ''
            df_gv_info = st.session_state.get('df_gv_info')
            if df_gv_info is not None and 'Tên_gv' in df_gv_info.columns and not df_gv_info.empty:
                ten_gv = str(df_gv_info['Tên_gv'].iloc[0]).strip()
            st.write(f"Tên giáo viên: {ten_gv}")
            # Lưu file tạm để openpyxl xử lý
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
                tmpfile.write(uploaded_tonghop_file.read())
                tmpfile_path = tmpfile.name
            wb = openpyxl.load_workbook(tmpfile_path)
            ws = wb.active
            # Tìm hàng theo tên giáo viên trong cột B (column 2)
            row_gv = None
            if ten_gv:
                for row in range(2, ws.max_row + 1):
                    cell_val = str(ws.cell(row=row, column=2).value).strip()
                    if cell_val == ten_gv:
                        row_gv = row
                        break
            if row_gv is not None:
                # Ghi tổng Tiết vào cột AB (28), tổng Tiết QĐ vào cột O (15)
                if tong_tiet is not None:
                    ws.cell(row=row_gv, column=28).value = tong_tiet/2
                if tong_tiet_qd is not None:
                    ws.cell(row=row_gv, column=15).value = tong_tiet_qd/2
                # Ghi tổng Tiết HK2 vào cột AC (29), tổng Tiết QĐ HK2 vào cột P (16)
                if tong_tiet_hk2 is not None:
                    ws.cell(row=row_gv, column=29).value = tong_tiet_hk2/2
                if tong_tiet_qd_hk2 is not None:
                    ws.cell(row=row_gv, column=16).value = tong_tiet_qd_hk2/2
                wb.save(tmpfile_path)
                with open(tmpfile_path, 'rb') as f:
                    st.download_button(
                        label="Tải xuống file tổng hợp đã cập nhật",
                        data=f.read(),
                        file_name="tonghop_capnhat.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                st.success(f"Đã cập nhật tổng Tiết ({tong_tiet/2}) vào cột AB, Tiết QĐ ({tong_tiet_qd/2}) vào cột O, tổng Tiết HK2 ({tong_tiet_hk2/2}) vào cột AC, Tiết QĐ HK2 ({tong_tiet_qd_hk2/2}) vào cột P cho giáo viên '{ten_gv}' trong file tổng hợp.")
            else:
                st.error(f"Không tìm thấy tên giáo viên '{ten_gv}' trong cột 'HỌ VÀ TÊN' của file tổng hợp.")
            import os
            os.unlink(tmpfile_path)
    # --- Mục 2: Cập nhật dữ liệu giáo viên vào file tổng hợp Google Sheet ---
    st.markdown("""
### 2. Cập nhật dữ liệu giáo viên vào file tổng hợp Google Sheet
""")
# Mục 1: Chọn mã khoa
ma_khoa = st.selectbox("Chọn mã khoa (1-9)", [str(i) for i in range(1, 10)], key="ma_khoa_google")

# Kết nối Google Drive và tìm các file Google Sheet có mã khoa đầu
on = st.toggle("Cập nhật tất cả")
try:
    import gspread
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds_dict = st.secrets["gcp_service_account"]
    folder_id = st.secrets["google_sheet"]["target_folder_id"]
    scopes = [
        "https://www.googleapis.com/auth/drive",
        "https://www.googleapis.com/auth/spreadsheets"
    ]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    gc = gspread.authorize(creds)
    drive_service = build('drive', 'v3', credentials=creds)
    # Tìm các file Google Sheet có tên bắt đầu bằng mã khoa
    query = f"'{folder_id}' in parents and mimeType='application/vnd.google-apps.spreadsheet' and name contains '{ma_khoa}'"
    results = drive_service.files().list(q=query, fields="files(id, name)").execute()
    sheets = results.get('files', [])
    sheet_names = [f"{s['name']} ({s['id']})" for s in sheets]
    if on:
        st.info("Đang cập nhật tất cả các sheet...")
        df_tonghop = []
        for sheet_info in sheet_names:
            ma_gv = str(sheet_info).split(' ')[0]
            sheet_id = sheet_info.split('(')[-1].replace(')', '')
            ten_gv_from_sheet = None
            if not df_giaovien.empty and 'Magv' in df_giaovien.columns and 'Tên giảng viên' in df_giaovien.columns:
                match_row = df_giaovien[df_giaovien['Magv'].astype(str) == ma_gv]
                if not match_row.empty:
                    ten_gv_from_sheet = match_row['Tên giảng viên'].iloc[0]
            # Đọc dữ liệu QĐ thừa từ Google Sheet
            qd_thua_sheet_hk1, qd_thua_sheet_hk2 = None, None
            try:
                sh = gc.open_by_key(sheet_id)
                # HK1
                try:
                    ws_giangday = sh.worksheet('output_giangday')
                    records_giangday = ws_giangday.get_all_records()
                    df_giangday_gsheet = pd.DataFrame(records_giangday)
                    if 'QĐ thừa' in df_giangday_gsheet.columns:
                        qd_thua_sheet_hk1 = df_giangday_gsheet['QĐ thừa'].sum()
                except Exception as e:
                    st.write(f"Không tìm thấy hoặc lỗi worksheet output_giangday: {e}")
                # HK2
                try:
                    ws_giangday_hk2 = sh.worksheet('output_giangday(HK2)')
                    records_giangday_hk2 = ws_giangday_hk2.get_all_records()
                    df_giangday_gsheet_hk2 = pd.DataFrame(records_giangday_hk2)
                    if 'QĐ thừa' in df_giangday_gsheet_hk2.columns:
                        qd_thua_sheet_hk2 = df_giangday_gsheet_hk2['QĐ thừa'].sum()
                except Exception as e:
                    st.write(f"Không tìm thấy hoặc lỗi worksheet output_giangday(HK2): {e}")
            except Exception as e:
                st.write(f"Lỗi khi đọc Google Sheet {sheet_info}: {e}")
            # Lưu vào danh sách tổng hợp
            df_tonghop.append({
                'ten_gv': ten_gv_from_sheet,
                'QĐ_thừa_hk1': qd_thua_sheet_hk1,
                'QĐ_thừa_hk2': qd_thua_sheet_hk2
            })
        df_tonghop = pd.DataFrame(df_tonghop)
        st.dataframe(df_tonghop)
        # Sau này, khi cập nhật file Excel tổng hợp giáo viên, dùng df_tonghop để ghi vào đúng dòng/cột
    else:
        df_tonghop = []
        # Thêm selectbox chọn sheet Google nếu chưa có
        selected_sheet = st.selectbox("Chọn Google Sheet cần xử lý:", sheet_names, key="select_gsheet_single") if sheet_names else None
        if selected_sheet:
                ma_gv = str(selected_sheet).split(' ')[0]
                sheet_id = selected_sheet.split('(')[-1].replace(')', '')
                ten_gv_from_sheet = None
                if not df_giaovien.empty and 'Magv' in df_giaovien.columns and 'Tên giảng viên' in df_giaovien.columns:
                    match_row = df_giaovien[df_giaovien['Magv'].astype(str) == ma_gv]
                    if not match_row.empty:
                        ten_gv_from_sheet = match_row['Tên giảng viên'].iloc[0]
                qd_thua_sheet_hk1, qd_thua_sheet_hk2 = None, None
                try:
                    sh = gc.open_by_key(sheet_id)
                    # HK1
                    try:
                        ws_giangday = sh.worksheet('output_giangday')
                        records_giangday = ws_giangday.get_all_records()
                        df_giangday_gsheet = pd.DataFrame(records_giangday)
                        if 'QĐ thừa' in df_giangday_gsheet.columns:
                            qd_thua_sheet_hk1 = df_giangday_gsheet['QĐ thừa'].sum()
                    except Exception as e:
                        st.write(f"Không tìm thấy hoặc lỗi worksheet output_giangday: {e}")
                    # HK2
                    try:
                        ws_giangday_hk2 = sh.worksheet('output_giangday(HK2)')
                        records_giangday_hk2 = ws_giangday_hk2.get_all_records()
                        df_giangday_gsheet_hk2 = pd.DataFrame(records_giangday_hk2)
                        if 'QĐ thừa' in df_giangday_gsheet_hk2.columns:
                            qd_thua_sheet_hk2 = df_giangday_gsheet_hk2['QĐ thừa'].sum()
                    except Exception as e:
                        st.write(f"Không tìm thấy hoặc lỗi worksheet output_giangday(HK2): {e}")
                except Exception as e:
                    st.write(f"Lỗi khi đọc Google Sheet {selected_sheet}: {e}")
                df_tonghop.append({
                    'ten_gv': ten_gv_from_sheet,
                    'QĐ_thừa_hk1': qd_thua_sheet_hk1,
                    'QĐ_thừa_hk2': qd_thua_sheet_hk2
                })
        df_tonghop = pd.DataFrame(df_tonghop)
        st.dataframe(df_tonghop)
except Exception as e:
    st.error(f"Lỗi kết nối Google API: {e}")

# Mục 2: Upload file excel có các cột như hình
uploaded_excel_gv = st.file_uploader("Upload file Excel tổng hợp giáo viên (cột như hình)", type=["xls", "xlsx"], key="excel_gv_google")
df_excel_gv = None

# --- Sau khi cập nhật vào file Excel upload, cho phép tải lại file đã cập nhật ---
if uploaded_excel_gv is not None and st.button("Cập nhật và tải lại file Excel đã upload"):
    import openpyxl
    import tempfile
    import io
    file_bytes = uploaded_excel_gv.read()  # Đọc một lần
    # Đọc bằng pandas
    try:
        df_excel_gv = pd.read_excel(io.BytesIO(file_bytes))
        #st.write("Đã đọc file Excel bằng pandas:")
        #st.dataframe(df_excel_gv.head())
    except Exception as e:
        st.write(f"Lỗi khi đọc file Excel bằng pandas: {e}")
        df_excel_gv = None
    # Đọc bằng openpyxl từ file_bytes
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
        tmpfile.write(file_bytes)
        tmpfile_path = tmpfile.name
    try:
        wb = openpyxl.load_workbook(tmpfile_path)
        ws = wb.active
        st.write(f"Số dòng sheet: {ws.max_row}, Số cột: {ws.max_column}")
        # Chỉ xử lý từ dòng 9 trở đi, cột 1 đến 27 (A đến AA)
        start_row = 9
        end_col = 27
        st.write(f"Chỉ xử lý từ dòng {start_row} đến hết, cột 1 đến {end_col} (A-AA)")
        # Duyệt từng giáo viên trong file excel, ghi dữ liệu từ df_tonghop
        if df_excel_gv is not None and not df_tonghop.empty:
            for idx, row in df_excel_gv.iterrows():
                # Lấy tên giáo viên từ cột Unnamed: 1 (hoặc tên cột đúng)
                col_gv = None
                for col in row.index:
                    if 'Tên giảng viên' in col or 'Họ và TÊN' in col or 'Unnamed: 1' in col:
                        col_gv = col
                        break
                ten_gv_row = str(row[col_gv]).strip() if col_gv and pd.notna(row[col_gv]) else None
                # Tìm thông tin QĐ thừa từ df_tonghop
                if ten_gv_row:
                    match = df_tonghop[df_tonghop['ten_gv'] == ten_gv_row]
                    if not match.empty:
                        qd_thua_hk1 = match['QĐ_thừa_hk1'].iloc[0]
                        qd_thua_hk2 = match['QĐ_thừa_hk2'].iloc[0]
                        # Tìm dòng trong file Excel upload, chỉ từ dòng 9 trở đi
                        for r in range(start_row, ws.max_row + 1):
                            cell_val = str(ws.cell(row=r, column=2).value).strip() if ws.cell(row=r, column=2).value else ""
                            if cell_val == ten_gv_row:
                                # Ghi QĐ thừa HK1 vào cột 15, HK2 vào cột 16
                                ws.cell(row=r, column=15).value = qd_thua_hk1
                                ws.cell(row=r, column=16).value = qd_thua_hk2
                                break
        wb.save(tmpfile_path)
        with open(tmpfile_path, 'rb') as f:
            st.download_button(
                label="Tải xuống file Excel đã cập nhật",
                data=f.read(),
                file_name="excel_gv_capnhat.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        st.write("Đã tạo nút tải file thành công.")
    except Exception as e:
        st.write(f"Lỗi khi xử lý/ghi file Excel: {e}")
    os.unlink(tmpfile_path)



